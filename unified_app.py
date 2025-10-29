"""
PDF Analyzer Unified - Integrates intelligent PDF analysis with advanced OCR
Combines:
- Automatic PDF type detection (textual/rasterized/hybrid)
- pdfplumber extraction with precise coordinates
- Advanced OCR with CV2 preprocessing and multi-PSM modes
- Number/dimension detection at 0° and 90°
- Density filtering and duplicate removal
"""

import os
from flask import Flask, render_template, request, jsonify, send_file, session
from werkzeug.utils import secure_filename
import pdfplumber
import fitz  # PyMuPDF
import pytesseract
from PIL import Image, ImageDraw
import cv2
import numpy as np
import base64
import io
import json
import csv
import datetime
from anthropic import Anthropic
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import uuid
from ai_providers import AIProviderManager

# Load environment variables from .env file
load_dotenv()

# Configure Tesseract path for Windows
if os.name == 'nt':
    tesseract_path = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    if os.path.exists(tesseract_path):
        pytesseract.pytesseract.tesseract_cmd = tesseract_path

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['TEMPLATES_FOLDER'] = 'saved_templates'
app.config['DIMENSION_PROMPTS_FOLDER'] = 'saved_dimension_prompts'
app.config['LAYOUT_PROMPTS_FOLDER'] = 'layout_prompts'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max
ALLOWED_EXTENSIONS = {'pdf'}

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['TEMPLATES_FOLDER'], exist_ok=True)
os.makedirs(app.config['DIMENSION_PROMPTS_FOLDER'], exist_ok=True)
os.makedirs(app.config['LAYOUT_PROMPTS_FOLDER'], exist_ok=True)

# Configure AI Provider Manager
DEMO_MODE = os.environ.get('DEMO_MODE', 'false').lower() == 'true'
ai_manager = AIProviderManager()

# Keep backward compatibility with legacy code
ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')
anthropic_client = None
if ANTHROPIC_API_KEY:
    anthropic_client = Anthropic(api_key=ANTHROPIC_API_KEY)

if DEMO_MODE:
    print("[WARNING] DEMO MODE ENABLED - Using simulated AI responses")
elif ai_manager.is_any_available():
    available = ai_manager.get_available_providers()
    print(f"AI Providers initialized: {', '.join(available.values())}")
    print(f"Current provider: {ai_manager.get_current_provider_name()}")
else:
    print("Warning: No AI providers configured. Add API keys to .env file.")


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ============================================================================
# PDF PROCESSING CLASS (from web_app.py)
# ============================================================================

class PDFProcessor:
    """Handles PDF analysis and text extraction with intelligent type detection"""

    def __init__(self, pdf_path):
        self.pdf_path = pdf_path
        self.pdf_type = None
        self.pages_info = []

    def detect_pdf_type(self):
        """Detect if PDF is textual, rasterized, or hybrid"""
        doc = fitz.open(self.pdf_path)
        text_pages = 0
        image_pages = 0

        for page_num in range(min(5, len(doc))):
            page = doc[page_num]
            text = page.get_text().strip()
            images = page.get_images()

            page_info = {
                'page_num': page_num,
                'has_text': len(text) > 100,
                'has_images': len(images) > 0,
                'text_length': len(text),
                'image_count': len(images)
            }

            if page_info['has_text']:
                text_pages += 1
            if page_info['has_images'] and not page_info['has_text']:
                image_pages += 1

            self.pages_info.append(page_info)

        doc.close()

        total_checked = len(self.pages_info)

        if text_pages == total_checked:
            self.pdf_type = 'textual'
        elif image_pages == total_checked:
            self.pdf_type = 'rasterized'
        else:
            self.pdf_type = 'hybrid'

        return self.pdf_type, self.pages_info

    def get_page_count(self):
        """Get total number of pages"""
        with pdfplumber.open(self.pdf_path) as pdf:
            return len(pdf.pages)

    def extract_with_pdfplumber(self, page_num=0, rotation=0, region=None):
        """Extract text with coordinates using pdfplumber"""
        extracted_data = []

        with pdfplumber.open(self.pdf_path) as pdf:
            if page_num >= len(pdf.pages):
                return extracted_data

            page = pdf.pages[page_num]

            if region:
                page = page.crop(region)

            words = page.extract_words(
                x_tolerance=3,
                y_tolerance=3,
                keep_blank_chars=False,
                use_text_flow=True
            )

            for word in words:
                extracted_data.append({
                    'text': word['text'],
                    'x0': word['x0'],
                    'y0': word['top'],
                    'x1': word['x1'],
                    'y1': word['bottom'],
                    'width': word['x1'] - word['x0'],
                    'height': word['bottom'] - word['top']
                })

        return extracted_data

    def get_full_text_pdfplumber(self):
        """Extract all text from all pages using pdfplumber"""
        full_text = []

        with pdfplumber.open(self.pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                text = page.extract_text()
                if text:
                    full_text.append(f"--- Page {page_num + 1} ---\n{text}\n")

        return "\n".join(full_text)

    def get_page_image(self, page_num=0, dpi=150):
        """Convert PDF page to base64 encoded image using PyMuPDF"""
        try:
            doc = fitz.open(self.pdf_path)

            if page_num >= len(doc):
                doc.close()
                return None

            page = doc[page_num]
            zoom = dpi / 72
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)

            img_data = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_data))
            doc.close()

            buffered = io.BytesIO()
            img.save(buffered, format="PNG")
            img_str = base64.b64encode(buffered.getvalue()).decode()
            return img_str

        except Exception as e:
            print(f"Error converting PDF page to image: {e}")
            return None

    def get_page_as_pil(self, page_num=0, dpi=300):
        """Get PDF page as PIL Image for OCR processing"""
        try:
            doc = fitz.open(self.pdf_path)

            if page_num >= len(doc):
                doc.close()
                return None

            page = doc[page_num]
            zoom = dpi / 72
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)

            img_data = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_data))
            doc.close()

            return img

        except Exception as e:
            print(f"Error getting PDF page as PIL: {e}")
            return None


# ============================================================================
# ADVANCED OCR FUNCTIONS (from app.py)
# ============================================================================

def preprocess_image(image):
    """Pre-process image to improve OCR using CV2"""
    img_array = np.array(image)
    gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)

    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
    enhanced = clahe.apply(gray)

    denoised = cv2.fastNlMeansDenoising(enhanced, None, h=5, templateWindowSize=7, searchWindowSize=21)
    _, binary = cv2.threshold(denoised, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    if np.mean(binary) > 127:
        binary = cv2.bitwise_not(binary)

    return Image.fromarray(binary)


def contains_numbers(text):
    """Check if text contains at least one number"""
    if not text:
        return False
    return any(c.isdigit() for c in text)

def is_date(text):
    """Verifica se il testo sembra una data"""
    import re
    # Patterns per date comuni: DD/MM/YYYY, DD-MM-YYYY, DD.MM.YYYY, YYYY-MM-DD, etc.
    date_patterns = [
        r'\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}',  # DD/MM/YYYY o simili
        r'\d{4}[/\-\.]\d{1,2}[/\-\.]\d{1,2}',      # YYYY-MM-DD o simili
        r'\d{1,2}\s+(?:gen|feb|mar|apr|mag|giu|lug|ago|set|ott|nov|dic)[a-z]*\s+\d{2,4}',  # DD mese YYYY
    ]
    for pattern in date_patterns:
        if re.search(pattern, text.lower()):
            return True
    return False

def is_reference(text):
    """Verifica se il testo sembra un riferimento progetto/documento"""
    import re
    # Patterns per riferimenti: numero progetto, codice documento, etc.
    # Es: "P-12345", "DOC-001", "REV.2", "N°123", ecc.
    ref_patterns = [
        r'(?:prog|project|doc|ref|cod|rif|rev|n)[°\.]?\s*[:\-]?\s*\d+',
        r'[A-Z]{2,}-\d+',  # Codici tipo "AB-123"
        r'(?:rev|version|ver)[\.:]?\s*\d+',
    ]
    for pattern in ref_patterns:
        if re.search(pattern, text.lower()):
            return True
    return False

def is_measurement_unit(text):
    """Verifica se il testo contiene unità di misura"""
    units = ['mm', 'cm', 'm', 'km', 'kg', 'g', 'l', 'ml', '°', 'kw', 'kva', 'v', 'a', 'hz', 'bar', 'mpa']
    text_lower = text.lower()
    return any(unit in text_lower for unit in units)

def extract_context_around(word, all_words, max_distance=50):
    """Estrae parole di contesto vicine a una parola data"""
    context = []
    word_center_x = word['x0'] + (word['x1'] - word['x0']) / 2
    word_center_y = word['top'] + (word['bottom'] - word['top']) / 2

    for other in all_words:
        if other['text'] == word['text'] and abs(other['x0'] - word['x0']) < 1:
            continue  # Stessa parola

        other_center_x = other['x0'] + (other['x1'] - other['x0']) / 2
        other_center_y = other['top'] + (other['bottom'] - other['top']) / 2

        distance = ((word_center_x - other_center_x)**2 + (word_center_y - other_center_y)**2)**0.5

        if distance < max_distance:
            context.append({
                'text': other['text'],
                'distance': distance,
                'direction': 'left' if other_center_x < word_center_x else 'right'
            })

    # Ordina per distanza
    context.sort(key=lambda x: x['distance'])
    return context[:5]  # Ritorna al massimo 5 parole di contesto

def reconstruct_rotated_words(chars):
    """
    Ricostruisce parole da caratteri ruotati a 90°/270°.
    I caratteri ruotati sono già orientati, quindi dobbiamo solo raggrupparli.
    """
    if len(chars) == 0:
        return []

    # Ordina caratteri per posizione (prima per x poi per y, dato che sono ruotati)
    sorted_chars = sorted(chars, key=lambda c: (c['x0'], c['y0']))

    words = []
    current_word_chars = []
    last_char = None

    for char in sorted_chars:
        if last_char is None:
            current_word_chars = [char]
            last_char = char
            continue

        # Calcola distanza dal carattere precedente
        x_dist = abs(char['x0'] - last_char['x0'])
        y_dist = abs(char['y0'] - last_char['y0'])

        # I caratteri ruotati sono adiacenti se hanno x simile e y vicino
        # (o viceversa a seconda della direzione di rotazione)
        is_adjacent = (x_dist < 15 and y_dist < 20)

        if is_adjacent:
            current_word_chars.append(char)
        else:
            # Salva la parola corrente
            if len(current_word_chars) > 0:
                word = merge_chars_to_word(current_word_chars)
                if word:
                    words.append(word)
            current_word_chars = [char]

        last_char = char

    # Salva l'ultima parola
    if len(current_word_chars) > 0:
        word = merge_chars_to_word(current_word_chars)
        if word:
            words.append(word)

    return words


def merge_chars_to_word(chars):
    """Unisce una lista di caratteri in una singola parola"""
    if len(chars) == 0:
        return None

    text = ''.join([c['text'] for c in chars])
    min_x = min(c['x0'] for c in chars)
    max_x = max(c['x1'] for c in chars)
    min_y = min(c['top'] for c in chars)
    max_y = max(c['bottom'] for c in chars)

    return {
        'text': text,
        'x0': min_x,
        'x1': max_x,
        'top': min_y,
        'bottom': max_y
    }


def merge_horizontal_digits(words):
    """
    Unisce cifre consecutive sulla stessa riga orizzontale.
    Questo cattura numeri come "1 2 3 4" che sono scritti con spazi larghi.
    """
    if len(words) == 0:
        return []

    merged = []
    used = set()

    # Ordina per riga (y) poi per colonna (x)
    sorted_words = sorted(enumerate(words), key=lambda iw: (iw[1]['top'], iw[1]['x0']))

    for idx_i, (i, word) in enumerate(sorted_words):
        if i in used:
            continue

        # Solo per caratteri singoli numerici
        if len(word['text'].strip()) != 1 or not word['text'].strip().isdigit():
            merged.append(word)
            used.add(i)
            continue

        # Trova altri numeri sulla stessa riga
        candidates = [i]
        base_y = word['top']
        base_x_end = word['x1']

        # Cerca numeri successivi sulla stessa riga
        for idx_j in range(idx_i + 1, len(sorted_words)):
            j, other = sorted_words[idx_j]

            if j in used:
                continue

            # Deve essere una singola cifra
            if len(other['text'].strip()) != 1 or not other['text'].strip().isdigit():
                continue

            # Controlla se è sulla stessa riga
            y_diff = abs(other['top'] - base_y)
            if y_diff > 5:  # Tolleranza 5px per la stessa riga
                break  # Siamo passati ad un'altra riga

            # Controlla distanza orizzontale (max 200px tra cifre)
            x_dist = other['x0'] - base_x_end
            if x_dist < 0 or x_dist > 200:
                continue

            candidates.append(j)
            base_x_end = other['x1']

        if len(candidates) >= 2:  # Almeno 2 cifre consecutive
            # Ordina per posizione x
            candidates.sort(key=lambda idx: words[idx]['x0'])

            # Unisci le cifre
            merged_text = ''.join([words[idx]['text'].strip() for idx in candidates])

            # Calcola bounding box unificata
            all_words = [words[idx] for idx in candidates]
            min_x = min(w['x0'] for w in all_words)
            max_x = max(w['x1'] for w in all_words)
            min_y = min(w['top'] for w in all_words)
            max_y = max(w['bottom'] for w in all_words)

            merged_word = {
                'text': merged_text,
                'x0': min_x,
                'x1': max_x,
                'top': min_y,
                'bottom': max_y
            }

            print(f"Merged horizontal digits: '{merged_text}' from {len(candidates)} digits")
            merged.append(merged_word)
            used.update(candidates)
        else:
            merged.append(word)
            used.add(i)

    return merged


def merge_vertical_chars(words):
    """
    Unisce caratteri verticali in parole/numeri completi.
    Gestisce due casi:
    1. Caratteri singoli con bbox verticale (altezza >> larghezza)
    2. Caratteri normali disposti verticalmente uno sotto l'altro
    """
    if len(words) == 0:
        return []

    merged = []
    used = set()

    for i, word in enumerate(words):
        if i in used:
            continue

        # Caso 1: Carattere singolo con bounding box verticale
        width = word['x1'] - word['x0']
        height = word['bottom'] - word['top']
        is_vertical_bbox = height > width * 2

        # Caso 2: Carattere singolo che potrebbe essere parte di testo verticale
        is_single_char = len(word['text'].strip()) == 1

        if not (is_vertical_bbox or is_single_char):
            # Parola già completa, mantieni così
            merged.append(word)
            used.add(i)
            continue

        # Cerca altri caratteri nella stessa "colonna verticale"
        candidates = [i]
        base_x = (word['x0'] + word['x1']) / 2
        base_y = word['top']
        base_height = height

        for j, other in enumerate(words):
            if j == i or j in used:
                continue

            # Considera solo caratteri singoli o con bbox verticale
            other_width = other['x1'] - other['x0']
            other_height = other['bottom'] - other['top']
            other_is_vertical_bbox = other_height > other_width * 2
            other_is_single_char = len(other['text'].strip()) == 1

            if not (other_is_vertical_bbox or other_is_single_char):
                continue

            other_x = (other['x0'] + other['x1']) / 2
            other_y = other['top']

            # Stessa colonna verticale (x simile)
            x_diff = abs(base_x - other_x)

            # Distanza verticale
            if other_y > base_y:
                y_gap = other_y - (word['bottom'] if j > i else word['top'])
            else:
                continue  # Salta caratteri sopra (li prenderemo quando sarà il loro turno)

            # Tolleranze più ampie per catturare testo ruotato
            x_tolerance = max(20, max(width, other_width) * 1.5)
            y_tolerance = max(50, base_height * 2)

            if x_diff < x_tolerance and y_gap < y_tolerance:
                candidates.append(j)

        if len(candidates) > 1:
            # Ordina per posizione verticale (top to bottom)
            candidates.sort(key=lambda idx: words[idx]['top'])

            # Unisci i caratteri
            merged_text = ''.join([words[idx]['text'].strip() for idx in candidates])

            # Calcola bounding box unificata
            all_words = [words[idx] for idx in candidates]
            min_x = min(w['x0'] for w in all_words)
            max_x = max(w['x1'] for w in all_words)
            min_y = min(w['top'] for w in all_words)
            max_y = max(w['bottom'] for w in all_words)

            merged_word = {
                'text': merged_text,
                'x0': min_x,
                'x1': max_x,
                'top': min_y,
                'bottom': max_y
            }

            print(f"Merged vertical text: '{merged_text}' from {len(candidates)} characters")
            merged.append(merged_word)
            used.update(candidates)
        else:
            merged.append(word)
            used.add(i)

    return merged

def extract_numbers_from_pdfplumber(pdf_path, page_num=0):
    """Estrae numeri, date e riferimenti da PDF testuale usando pdfplumber con coordinate"""
    results = []

    with pdfplumber.open(pdf_path) as pdf:
        if page_num >= len(pdf.pages):
            return results

        page = pdf.pages[page_num]

        # FASE 1: Estrai caratteri con informazioni di rotazione
        print(f"\n=== PDFPLUMBER CHAR-LEVEL EXTRACTION ===")
        chars = page.chars
        print(f"Total chars extracted: {len(chars)}")

        # Raggruppa caratteri ruotati (90° o 270°)
        rotated_chars = []
        normal_chars = []

        for char in chars:
            # pdfplumber usa "matrix" per la trasformazione
            # matrix[0] e matrix[3] indicano la rotazione
            # Rotazione 90°: matrix = (0, 1, -1, 0, x, y)
            # Rotazione 0°: matrix = (1, 0, 0, 1, x, y)
            matrix = char.get('matrix', (1, 0, 0, 1, 0, 0))

            # Rileva rotazione
            is_rotated_90 = abs(matrix[0]) < 0.1 and abs(matrix[1]) > 0.9  # 90° clockwise
            is_rotated_270 = abs(matrix[0]) < 0.1 and abs(matrix[3]) < -0.9  # 270° clockwise (90° counter)

            if is_rotated_90 or is_rotated_270:
                rotated_chars.append(char)
            else:
                normal_chars.append(char)

        print(f"Rotated chars: {len(rotated_chars)}, Normal chars: {len(normal_chars)}")

        # FASE 2: Estrai words normalmente
        words = page.extract_words(
            x_tolerance=3,
            y_tolerance=3,
            keep_blank_chars=False,
            use_text_flow=True
        )

        print(f"\n=== WORD-LEVEL EXTRACTION ===")
        print(f"Total words extracted: {len(words)}")

        # Stampa alcuni esempi di caratteri singoli per debug
        single_chars = [w for w in words if len(w['text'].strip()) == 1]
        print(f"Single characters found: {len(single_chars)}")

        # FASE 3A: Unisci caratteri orizzontali (numeri sulla stessa riga)
        print(f"\nAttempting to merge horizontal digit sequences...")
        words_before_h = len(words)
        words = merge_horizontal_digits(words)
        words_after_h = len(words)
        print(f"Words before horizontal merge: {words_before_h}, after merge: {words_after_h}, reduced by: {words_before_h - words_after_h}")

        # FASE 3B: Unisci caratteri verticali (da words)
        print(f"\nAttempting to merge vertical characters...")
        words_before = len(words)
        words = merge_vertical_chars(words)
        words_after = len(words)
        print(f"Words before merge: {words_before}, after merge: {words_after}, reduced by: {words_before - words_after}")

        # FASE 4: Ricostruisci parole da caratteri ruotati
        if len(rotated_chars) > 0:
            print(f"\nReconstructing words from {len(rotated_chars)} rotated characters...")
            rotated_words = reconstruct_rotated_words(rotated_chars)
            print(f"Created {len(rotated_words)} words from rotated chars")
            words.extend(rotated_words)

        for idx, word in enumerate(words):
            text = word['text'].strip()
            if not text:
                continue

            # Determina il tipo di dato
            data_type = 'text'
            if contains_numbers(text):
                data_type = 'number'
            if is_date(text):
                data_type = 'date'
            if is_reference(text):
                data_type = 'reference'
            if is_measurement_unit(text):
                data_type = 'unit'

            # Estrai solo dati rilevanti (numeri, date, riferimenti, unità)
            if data_type in ['number', 'date', 'reference', 'unit']:
                # Estrai contesto (parole vicine)
                context = extract_context_around(word, words, max_distance=50)
                context_text = ' '.join([c['text'] for c in context[:3]])

                results.append({
                    'id': idx,
                    'text': text,
                    'type': data_type,
                    'bbox': {
                        'x': word['x0'],
                        'y': word['top'],
                        'width': word['x1'] - word['x0'],
                        'height': word['bottom'] - word['top']
                    },
                    'context': context_text,
                    'source': 'pdfplumber',
                    'confidence': 100  # pdfplumber è preciso al 100%
                })

    return results


def is_horizontal_text(bbox, text):
    """Verifica se il testo è orizzontale in base al rapporto larghezza/altezza"""
    width = bbox['width']
    height = bbox['height']

    if len(text) == 1:
        return height <= width * 3
    else:
        return width >= height * 0.6

def merge_number_x_number(results):
    """Unisce pattern come '25', 'x', '30' in '25x30' (con o senza spazi)"""
    if len(results) < 2:
        return results

    merged = []
    i = 0

    while i < len(results):
        current = results[i]
        current_is_num = current['text'].replace('.', '').replace(',', '').replace('-', '').isdigit()

        if current_is_num and i < len(results) - 1:
            found_pattern = False
            j = i + 1

            while j < min(len(results), i + 6):
                candidate_x = results[j]

                if candidate_x['text'].lower() == 'x':
                    for k in range(j + 1, min(len(results), j + 4)):
                        candidate_num = results[k]
                        next_is_num = candidate_num['text'].replace('.', '').replace(',', '').replace('-', '').isdigit()

                        if next_is_num:
                            y_tolerance = max(current['bbox']['height'], candidate_num['bbox']['height']) * 0.5
                            same_line = abs(current['bbox']['y'] - candidate_x['bbox']['y']) < y_tolerance and \
                                       abs(candidate_x['bbox']['y'] - candidate_num['bbox']['y']) < y_tolerance

                            total_gap = candidate_num['bbox']['x'] - (current['bbox']['x'] + current['bbox']['width'])
                            close_together = total_gap < 100

                            if same_line and close_together:
                                merged_text = f"{current['text']}x{candidate_num['text']}"

                                all_elements = [current] + results[i+1:k+1]

                                min_x = min(el['bbox']['x'] for el in all_elements)
                                max_x = max(el['bbox']['x'] + el['bbox']['width'] for el in all_elements)
                                min_y = min(el['bbox']['y'] for el in all_elements)
                                max_y = max(el['bbox']['y'] + el['bbox']['height'] for el in all_elements)

                                merged_result = {
                                    'text': merged_text,
                                    'bbox': {
                                        'x': min_x,
                                        'y': min_y,
                                        'width': max_x - min_x,
                                        'height': max_y - min_y
                                    },
                                    'confidence': max(el['confidence'] for el in all_elements)
                                }

                                merged.append(merged_result)
                                i = k + 1
                                found_pattern = True
                                break

                    if found_pattern:
                        break

                j += 1

            if found_pattern:
                continue

        merged.append(results[i])
        i += 1

    return merged

def process_single_rotation(image, angle, label, min_conf=60):
    """Processa l'immagine con una specifica rotazione"""
    print(f"\n=== Elaborazione {label} (soglia: {min_conf}%) ===")

    if angle != 0:
        rotated_image = image.rotate(-angle, expand=True)
    else:
        rotated_image = image

    processed_image = preprocess_image(rotated_image)

    all_results = []

    # PSM 6 - Blocco uniforme
    print("Esecuzione OCR con PSM 6...")
    config1 = r'--oem 3 --psm 6'
    ocr_data1 = pytesseract.image_to_data(processed_image, lang='ita+eng',
                                          config=config1,
                                          output_type=pytesseract.Output.DICT)
    all_results.extend(parse_ocr_data(ocr_data1, min_conf=min_conf))

    # PSM 11 - Testo sparso
    print("Esecuzione OCR con PSM 11...")
    config2 = r'--oem 3 --psm 11'
    ocr_data2 = pytesseract.image_to_data(processed_image, lang='ita+eng',
                                          config=config2,
                                          output_type=pytesseract.Output.DICT)
    all_results.extend(parse_ocr_data(ocr_data2, min_conf=min_conf))

    # PSM 3 - Auto page segmentation
    print("Esecuzione OCR con PSM 3...")
    config3 = r'--oem 3 --psm 3'
    ocr_data3 = pytesseract.image_to_data(processed_image, lang='ita+eng',
                                          config=config3,
                                          output_type=pytesseract.Output.DICT)
    all_results.extend(parse_ocr_data(ocr_data3, min_conf=min_conf))

    # Filtra solo elementi che contengono numeri o 'x'
    results_with_numbers = [r for r in all_results if contains_numbers(r['text']) or r['text'].lower() == 'x']

    # Filtra solo testo orizzontale
    horizontal_results = [r for r in results_with_numbers if is_horizontal_text(r['bbox'], r['text'])]

    print(f"Filtrati {len(results_with_numbers) - len(horizontal_results)} elementi verticali")

    # Rimuovi duplicati
    unique_results = remove_duplicates_simple(horizontal_results)

    # Ordina per posizione
    unique_results.sort(key=lambda r: (r['bbox']['y'], r['bbox']['x']))

    # Unisci pattern "numero x numero"
    merged_results = merge_number_x_number(unique_results)

    # Filtra di nuovo per rimuovere eventuali 'x' rimasti isolati
    final_results = [r for r in merged_results if contains_numbers(r['text'])]

    # Aggiungi ID
    for idx, result in enumerate(final_results):
        result['id'] = idx

    print(f"Trovati {len(final_results)} elementi con numeri per {label}")

    return final_results

def transform_bbox_from_90_to_0(bbox, rotated_width, rotated_height, original_width, original_height):
    """Trasforma le coordinate di un bbox dalla rotazione 90° a 0°"""
    x_rot = bbox['x']
    y_rot = bbox['y']
    w_rot = bbox['width']
    h_rot = bbox['height']

    x_orig = y_rot
    y_orig = original_height - x_rot - w_rot

    w_orig = h_rot
    h_orig = w_rot

    return {
        'x': x_orig,
        'y': y_orig,
        'width': w_orig,
        'height': h_orig
    }

def calculate_text_density_around(bbox, all_elements, radius_multiplier=3.0):
    """Calcola la densità di testo nell'area circostante un bounding box"""
    center_x = bbox['x'] + bbox['width'] / 2
    center_y = bbox['y'] + bbox['height'] / 2

    search_radius = max(bbox['width'], bbox['height']) * radius_multiplier

    nearby_count = 0
    for element in all_elements:
        elem_center_x = element['x'] + element['width'] / 2
        elem_center_y = element['y'] + element['height'] / 2

        distance = ((center_x - elem_center_x) ** 2 + (center_y - elem_center_y) ** 2) ** 0.5

        if distance < search_radius and distance > 0:
            nearby_count += 1

    return nearby_count

def remove_overlapping_rectangles(numbers_0deg, numbers_90deg):
    """Rimuove rettangoli sovrapposti, mantenendo quello con lunghezza maggiore"""
    filtered_0deg = []
    filtered_90deg = []

    for num_0 in numbers_0deg:
        bbox_0 = num_0['bbox']
        overlaps = False

        for num_90 in numbers_90deg:
            bbox_90 = num_90['bbox']

            x_overlap = max(0, min(bbox_0['x'] + bbox_0['width'], bbox_90['x'] + bbox_90['width']) -
                           max(bbox_0['x'], bbox_90['x']))
            y_overlap = max(0, min(bbox_0['y'] + bbox_0['height'], bbox_90['y'] + bbox_90['height']) -
                           max(bbox_0['y'], bbox_90['y']))

            overlap_area = x_overlap * y_overlap

            area_0 = bbox_0['width'] * bbox_0['height']
            area_90 = bbox_90['width'] * bbox_90['height']
            min_area = min(area_0, area_90)

            if min_area > 0 and overlap_area / min_area > 0.5:
                length_0 = bbox_0['width']
                length_90 = bbox_90['height']

                if length_90 > length_0:
                    overlaps = True
                    break

        if not overlaps:
            filtered_0deg.append(num_0)

    for num_90 in numbers_90deg:
        bbox_90 = num_90['bbox']
        overlaps = False

        for num_0 in filtered_0deg:
            bbox_0 = num_0['bbox']

            x_overlap = max(0, min(bbox_0['x'] + bbox_0['width'], bbox_90['x'] + bbox_90['width']) -
                           max(bbox_0['x'], bbox_90['x']))
            y_overlap = max(0, min(bbox_0['y'] + bbox_0['height'], bbox_90['y'] + bbox_90['height']) -
                           max(bbox_0['y'], bbox_90['y']))

            overlap_area = x_overlap * y_overlap

            area_0 = bbox_0['width'] * bbox_0['height']
            area_90 = bbox_90['width'] * bbox_90['height']
            min_area = min(area_0, area_90)

            if min_area > 0 and overlap_area / min_area > 0.5:
                length_0 = bbox_0['width']
                length_90 = bbox_90['height']

                if length_0 >= length_90:
                    overlaps = True
                    break

        if not overlaps:
            filtered_90deg.append(num_90)

    return filtered_0deg, filtered_90deg

def extract_numbers_advanced(image, min_conf=60):
    """
    Estrae testo contenente numeri dall'immagine a 0° e 90°, unificando i risultati
    Restituisce numeri con marker di provenienza (0deg/90deg) per colorazione
    """
    # Elaborazione a 0° (orizzontale)
    numbers_from_0 = process_single_rotation(image, 0, '0 gradi (Orizzontale)', min_conf=min_conf)

    # Elaborazione a 90° (verticale->orizzontale)
    numbers_from_90_raw = process_single_rotation(image, 90, '90 gradi (Verticale)', min_conf=min_conf)

    # Ottieni dimensioni
    original_width, original_height = image.size
    rotated_image = image.rotate(-90, expand=True)
    rotated_width, rotated_height = rotated_image.size

    # Filtra i numeri da 90° in base alla densità
    print("\nFiltraggio elementi 90° in aree ad alta densità...")
    processed_image = preprocess_image(rotated_image)
    config = r'--oem 3 --psm 6'
    ocr_data = pytesseract.image_to_data(processed_image, lang='ita+eng',
                                         config=config,
                                         output_type=pytesseract.Output.DICT)

    all_elements = []
    n_boxes = len(ocr_data['text'])
    for i in range(n_boxes):
        text = ocr_data['text'][i].strip()
        if not text:
            continue
        try:
            conf = int(ocr_data['conf'][i])
        except:
            continue
        if conf < 30:
            continue

        x = ocr_data['left'][i]
        y = ocr_data['top'][i]
        w = ocr_data['width'][i]
        h = ocr_data['height'][i]

        if w < 3 or h < 3:
            continue

        all_elements.append({'x': x, 'y': y, 'width': w, 'height': h})

    print(f"Trovati {len(all_elements)} elementi totali nell'immagine ruotata")

    filtered_90_results = []
    removed_count = 0
    density_threshold = 8

    for result in numbers_from_90_raw:
        density = calculate_text_density_around(result['bbox'], all_elements, radius_multiplier=2.5)

        if density <= density_threshold:
            filtered_90_results.append(result)
        else:
            removed_count += 1
            print(f"Rimosso '{result['text']}' (densità: {density})")

    print(f"Filtrati {removed_count} elementi da 90° in aree dense")
    print(f"Mantenuti {len(filtered_90_results)} elementi da 90°")

    # Trasforma coordinate 90° -> 0°
    numbers_from_90 = []
    for result in filtered_90_results:
        transformed_bbox = transform_bbox_from_90_to_0(
            result['bbox'],
            rotated_width,
            rotated_height,
            original_width,
            original_height
        )
        numbers_from_90.append({
            'text': result['text'],
            'bbox': transformed_bbox,
            'confidence': result['confidence'],
            'source': '90deg',
            'id': result['id']
        })

    # Aggiungi marker ai numeri da 0°
    for result in numbers_from_0:
        result['source'] = '0deg'

    # Rimuovi sovrapposizioni
    print(f"\nRimozione sovrapposizioni: {len(numbers_from_0)} blu + {len(numbers_from_90)} fucsia")
    numbers_from_0_filtered, numbers_from_90_filtered = remove_overlapping_rectangles(numbers_from_0, numbers_from_90)
    removed_0 = len(numbers_from_0) - len(numbers_from_0_filtered)
    removed_90 = len(numbers_from_90) - len(numbers_from_90_filtered)
    print(f"Rimossi: {removed_0} blu, {removed_90} fucsia")
    print(f"Mantenuti: {len(numbers_from_0_filtered)} blu + {len(numbers_from_90_filtered)} fucsia")

    # Unisci tutti i numeri
    all_numbers = numbers_from_0_filtered + numbers_from_90_filtered

    # Ri-assegna gli ID
    for idx, result in enumerate(all_numbers):
        result['id'] = idx

    print(f"Unificati: {len(numbers_from_0_filtered)} + {len(numbers_from_90_filtered)} = {len(all_numbers)} totali")

    return all_numbers, numbers_from_0_filtered, numbers_from_90_filtered


def parse_ocr_data(ocr_data, min_conf=30):
    """Parse OCR data and return list of results"""
    results = []
    n_boxes = len(ocr_data['text'])

    for i in range(n_boxes):
        text = ocr_data['text'][i].strip()
        if not text or len(text) < 1:
            continue

        try:
            conf = int(ocr_data['conf'][i])
        except:
            continue

        if conf < min_conf:
            continue

        x = ocr_data['left'][i]
        y = ocr_data['top'][i]
        w = ocr_data['width'][i]
        h = ocr_data['height'][i]

        if w < 5 or h < 5:
            continue

        results.append({
            'text': text,
            'bbox': {
                'x': x,
                'y': y,
                'width': w,
                'height': h
            },
            'confidence': conf
        })

    return results


def remove_duplicates_simple(results):
    """Remove simple duplicates based on overlap and text"""
    if not results:
        return []

    unique = []

    for result in results:
        is_dup = False

        for u in unique:
            if result['text'].lower() == u['text'].lower():
                b1 = result['bbox']
                b2 = u['bbox']

                x_overlap = max(0, min(b1['x']+b1['width'], b2['x']+b2['width']) - max(b1['x'], b2['x']))
                y_overlap = max(0, min(b1['y']+b1['height'], b2['y']+b2['height']) - max(b1['y'], b2['y']))

                overlap_area = x_overlap * y_overlap
                area1 = b1['width'] * b1['height']

                if area1 > 0 and overlap_area / area1 > 0.5:
                    is_dup = True
                    if result['confidence'] > u['confidence']:
                        u['confidence'] = result['confidence']
                        u['bbox'] = result['bbox']
                    break

        if not is_dup:
            unique.append(result)

    return unique


def draw_boxes_on_image(image, results, highlight_id=None):
    """Draw bounding boxes on image"""
    img_with_boxes = image.copy()
    draw = ImageDraw.Draw(img_with_boxes)

    for idx, result in enumerate(results):
        bbox = result['bbox']
        x, y, w, h = bbox['x'], bbox['y'], bbox['width'], bbox['height']

        result_id = result.get('id', idx)

        if result_id == highlight_id:
            color = '#FFD700'  # Gold for highlighted
            width = 5
        else:
            color = '#0066FF'  # Blue
            width = 3

        for i in range(width):
            draw.rectangle([x-i, y-i, x + w + i, y + h + i], outline=color, width=1)

    return img_with_boxes

def draw_unified_boxes(image, numbers_0deg, numbers_90deg, highlight_id=None):
    """Disegna rettangoli blu per 0° e fucsia per 90° sull'immagine a 0°"""
    img_with_boxes = image.copy()
    draw = ImageDraw.Draw(img_with_boxes)

    # Disegna rettangoli per numeri da 0° (blu)
    for result in numbers_0deg:
        bbox = result['bbox']
        x, y, w, h = bbox['x'], bbox['y'], bbox['width'], bbox['height']
        result_id = result.get('id')

        if result_id == highlight_id:
            color = '#FFD700'  # Oro per evidenziato
            width = 5
        else:
            color = '#0066FF'  # Blu per 0°
            width = 3

        for i in range(width):
            draw.rectangle([x-i, y-i, x + w + i, y + h + i], outline=color, width=1)

    # Disegna rettangoli per numeri da 90° (fucsia)
    for result in numbers_90deg:
        bbox = result['bbox']
        x, y, w, h = bbox['x'], bbox['y'], bbox['width'], bbox['height']
        result_id = result.get('id')

        if result_id == highlight_id:
            color = '#FFD700'  # Oro per evidenziato
            width = 5
        else:
            color = '#FF00FF'  # Fucsia per 90°
            width = 3

        for i in range(width):
            draw.rectangle([x-i, y-i, x + w + i, y + h + i], outline=color, width=1)

    return img_with_boxes

def draw_pdfplumber_boxes(image, dpi, results, highlight_id=None):
    """Disegna rettangoli colorati per tipo su immagine da pdfplumber"""
    img_with_boxes = image.copy()
    draw = ImageDraw.Draw(img_with_boxes)

    # Scala fattore: pdfplumber usa punti (72 DPI), immagine può essere a DPI diverso
    scale = dpi / 72.0

    # Colori per tipo
    type_colors = {
        'number': '#0066FF',      # Blu per numeri
        'date': '#FF6B35',        # Arancione per date
        'reference': '#9D4EDD',   # Viola per riferimenti
        'unit': '#06FFA5'         # Verde acqua per unità
    }

    for result in results:
        bbox = result['bbox']
        x = bbox['x'] * scale
        y = bbox['y'] * scale
        w = bbox['width'] * scale
        h = bbox['height'] * scale
        result_id = result.get('id')
        data_type = result.get('type', 'number')

        if result_id == highlight_id:
            color = '#FFD700'  # Oro per evidenziato
            width = 5
        else:
            color = type_colors.get(data_type, '#0066FF')
            width = 3

        for i in range(width):
            draw.rectangle([x-i, y-i, x + w + i, y + h + i], outline=color, width=1)

    return img_with_boxes


def image_to_base64(image):
    """Convert PIL image to base64"""
    buffered = io.BytesIO()
    image.save(buffered, format="PNG")
    img_str = base64.b64encode(buffered.getvalue()).decode()
    return f"data:image/png;base64,{img_str}"


# ============================================================================
# CLAUDE OPUS INTEGRATION FUNCTIONS
# ============================================================================

def analyze_text_with_opus(extracted_data, pdf_type):
    """
    Feature 1: Intelligent text analysis
    Analyzes extracted numbers, dates, references and provides insights
    """
    if DEMO_MODE:
        # Demo mode - return simulated response
        return {'success': True, 'analysis': {
            'numeri_chiave': ['Dimensioni principali: 1234x5678 (simulato)', 'Valore massimo: 9999 (simulato)'],
            'date_critiche': ['Data documento: 2024-01-15 (simulato)'],
            'riferimenti': ['REF-001 (simulato)', 'PROG-2024 (simulato)'],
            'anomalie': [],
            'pattern': ['Numeri sequenziali rilevati (simulato)'],
            'riepilogo': 'Analisi simulata in DEMO MODE. Configura una API key valida per usare Claude Opus reale.'
        }}

    # Get current AI provider
    provider = ai_manager.get_current_provider()
    if not provider:
        return {'error': 'No AI provider configured. Configure an API key in .env file.'}

    # Prepare data summary
    numbers_summary = []
    for item in extracted_data:
        numbers_summary.append({
            'text': item['text'],
            'type': item.get('type', 'unknown'),
            'context': item.get('context', '')
        })

    prompt = f"""Analizza i seguenti dati estratti da un PDF {pdf_type}:

{json.dumps(numbers_summary, indent=2, ensure_ascii=False)}

Fornisci un'analisi strutturata che includa:
1. **Numeri chiave**: Identifica le misure più importanti o ricorrenti
2. **Date critiche**: Evidenzia le date significative trovate
3. **Riferimenti**: Elenca i codici/riferimenti di progetto o documento
4. **Anomalie**: Segnala eventuali valori sospetti o incoerenze
5. **Pattern**: Identifica schemi o relazioni tra i dati

Rispondi in italiano in formato JSON con questa struttura:
{{
  "numeri_chiave": ["lista di numeri importanti con spiegazione"],
  "date_critiche": ["lista di date con contesto"],
  "riferimenti": ["lista di riferimenti trovati"],
  "anomalie": ["eventuali anomalie rilevate"],
  "pattern": ["pattern o relazioni identificate"],
  "riepilogo": "breve riepilogo generale"
}}"""

    try:
        # Use the current AI provider
        analysis_text = provider.analyze_text(prompt, "")

        # Try to parse as JSON
        try:
            analysis = json.loads(analysis_text)
        except:
            # If not valid JSON, return as text
            analysis = {'riepilogo': analysis_text}

        return {'success': True, 'analysis': analysis}

    except Exception as e:
        return {'error': f'Error calling AI provider ({provider.get_name()}): {str(e)}'}


def vision_guided_extraction(image_base64, pdf_type):
    """
    Feature 2: Vision-guided extraction
    Uses Opus vision capabilities to analyze PDF page images
    """
    if DEMO_MODE:
        # Demo mode - return simulated vision response
        return {'success': True, 'vision_analysis': '''ANALISI VISIONE SIMULATA (DEMO MODE):

📊 Numeri e Misure Identificate:
- Rilevati valori numerici nell'area centrale del documento
- Dimensioni principali visibili: circa 5-10 numeri prominenti
- Alcuni numeri sembrano essere misure tecniche (mm, cm)

📅 Date:
- Possibile data nell'intestazione (formato DD/MM/YYYY)

🔗 Elementi Grafici:
- Tabella o griglia strutturata visibile
- Bordi e linee di separazione presenti
- Layout professionale/tecnico

🔄 Testo Verticale/Ruotato:
- Alcuni elementi sembrano orientati verticalmente
- Possibili etichette laterali

⚠️ Nota: Questa è una risposta simulata. Per un'analisi visiva reale, configura una API key valida di Anthropic.'''}

    # Get current AI provider
    provider = ai_manager.get_current_provider()
    if not provider:
        return {'error': 'No AI provider configured. Configure an API key in .env file.'}

    prompt = f"""Analizza questa pagina di un PDF {pdf_type} e identifica:

1. **Numeri e misure**: Tutti i valori numerici visibili con il loro contesto
2. **Date**: Eventuali date presenti
3. **Riferimenti**: Codici, numeri di progetto o documento
4. **Elementi grafici**: Diagrammi, tabelle, schemi che contengono informazioni importanti
5. **Testo verticale o ruotato**: Qualsiasi testo non orizzontale

Fornisci un'analisi dettagliata in italiano, evidenziando elementi che potrebbero essere difficili da estrarre con OCR tradizionale."""

    try:
        # Use the current AI provider's vision capability
        vision_analysis = provider.analyze_vision(prompt, image_base64)
        return {'success': True, 'vision_analysis': vision_analysis}

    except Exception as e:
        return {'error': f'Error calling AI provider vision ({provider.get_name()}): {str(e)}'}


def answer_question_about_pdf(question, full_text, extracted_data):
    """
    Feature 3: Question-Answering
    Allows users to ask questions about the PDF content
    """
    if DEMO_MODE:
        # Demo mode - return simulated answer
        return {'success': True, 'answer': f'''RISPOSTA SIMULATA (DEMO MODE):

Domanda ricevuta: "{question}"

Basandomi sul documento analizzato, posso fornire questa risposta di esempio:

- Ho identificato diversi numeri rilevanti nel documento
- Sono presenti misure tecniche e dimensioni
- Il documento sembra contenere informazioni strutturate

⚠️ Nota: Questa è una risposta simulata generica. Per risposte precise basate sul contenuto reale del PDF, configura una API key valida di Anthropic.

Per ottenere risposte accurate, attiva l'API di Claude Opus sul tuo account Anthropic.'''}

    # Get current AI provider
    provider = ai_manager.get_current_provider()
    if not provider:
        return {'error': 'No AI provider configured. Configure an API key in .env file.'}

    # Create context from full text and extracted data
    context = f"""Testo completo del PDF:
{full_text[:5000]}  # Limit to avoid token limits

Dati estratti (numeri, date, riferimenti):
{json.dumps(extracted_data[:50], indent=2, ensure_ascii=False)}  # First 50 items
"""

    prompt = f"""Basandoti sul seguente contenuto di un PDF, rispondi alla domanda dell'utente in modo preciso e dettagliato.

{context}

Domanda: {question}

Rispondi in italiano, citando i dati specifici dal documento quando possibile."""

    try:
        # Use the current AI provider's chat capability
        answer = provider.chat([{"role": "user", "content": prompt}])
        return {'success': True, 'answer': answer}

    except Exception as e:
        return {'error': f'Error calling AI provider ({provider.get_name()}): {str(e)}'}


def summarize_document(full_text, extracted_data, pdf_type):
    """
    Feature 4: Automatic document summarization
    Generates structured summary of entire PDF
    """
    if DEMO_MODE:
        # Demo mode - return simulated summary
        return {'success': True, 'summary': {
            'tipo_documento': 'Documento tecnico/disegno (simulato)',
            'scopo': 'Fornire specifiche tecniche e dimensioni (simulato)',
            'informazioni_chiave': [
                'Contiene misure e dimensioni tecniche',
                'Include riferimenti a standard o progetti',
                'Presenta layout strutturato con tabelle/griglie'
            ],
            'numeri_rilevanti': [
                'Dimensioni principali rilevate nel documento',
                'Valori numerici tecnici presenti'
            ],
            'date_importanti': ['Date di emissione/revisione (se presenti)'],
            'riferimenti': ['Codici progetto o riferimenti standard'],
            'conclusioni': 'DEMO MODE ATTIVO - Questo e\' un riepilogo simulato. Per analisi reale del documento, configura una API key valida di Anthropic Claude Opus.'
        }}

    # Get current AI provider
    provider = ai_manager.get_current_provider()
    if not provider:
        return {'error': 'No AI provider configured. Configure an API key in .env file.'}

    # Prepare summary of extracted data by type
    data_by_type = {}
    for item in extracted_data:
        item_type = item.get('type', 'unknown')
        if item_type not in data_by_type:
            data_by_type[item_type] = []
        data_by_type[item_type].append(item['text'])

    prompt = f"""Analizza questo documento PDF {pdf_type} e crea un riepilogo strutturato.

Testo del documento (prime 6000 caratteri):
{full_text[:6000]}

Dati estratti per tipo:
{json.dumps(data_by_type, indent=2, ensure_ascii=False)}

Crea un riepilogo strutturato in formato JSON con:
{{
  "tipo_documento": "descrizione del tipo di documento",
  "scopo": "scopo principale del documento",
  "informazioni_chiave": ["lista di informazioni più importanti"],
  "numeri_rilevanti": ["numeri/misure principali con contesto"],
  "date_importanti": ["date significative"],
  "riferimenti": ["codici/riferimenti identificati"],
  "conclusioni": "sintesi finale del contenuto"
}}

Rispondi SOLO con il JSON, senza testo aggiuntivo."""

    try:
        # Use the current AI provider
        summary_text = provider.analyze_text(prompt, "")

        # Try to parse as JSON
        try:
            summary = json.loads(summary_text)
        except:
            # If not valid JSON, return as text
            summary = {'conclusioni': summary_text}

        return {'success': True, 'summary': summary}

    except Exception as e:
        return {'error': f'Error calling AI provider ({provider.get_name()}): {str(e)}'}


# ============================================================================
# FLASK ROUTES
# ============================================================================

@app.route('/')
def index():
    # Leggi la versione dal file VERSION.txt
    version = "0.0"
    try:
        with open('VERSION.txt', 'r') as f:
            version = f.read().strip()
    except Exception as e:
        print(f"Errore lettura VERSION.txt: {e}")
    return render_template('unified.html', version=version)


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'current.pdf')
        file.save(filepath)

        processor = PDFProcessor(filepath)
        pdf_type, pages_info = processor.detect_pdf_type()
        page_count = processor.get_page_count()

        # Extract full text
        if pdf_type in ['textual', 'hybrid']:
            full_text = processor.get_full_text_pdfplumber()
        else:
            full_text = "PDF rasterizzato - usa OCR avanzato per l'estrazione"

        # Get first page as PIL image for processing
        image = processor.get_page_as_pil(page_num=0, dpi=300)

        # Salva l'immagine originale
        original_path = os.path.join(app.config['UPLOAD_FOLDER'], 'original.png')
        image.save(original_path)

        # Strategia di estrazione basata sul tipo di PDF
        try:
            if pdf_type in ['textual', 'hybrid']:
                # Usa pdfplumber per PDF testuali
                print(f"PDF {pdf_type} rilevato - Uso pdfplumber per estrazione")
                all_numbers = extract_numbers_from_pdfplumber(filepath, page_num=0)

                # Disegna rettangoli colorati per tipo
                img_with_boxes = draw_pdfplumber_boxes(image, 300, all_numbers)
                page_image = image_to_base64(img_with_boxes)

                # Salva i risultati
                results_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ocr_results.json')
                with open(results_path, 'w', encoding='utf-8') as f:
                    json.dump({
                        'all_numbers': all_numbers,
                        'extraction_method': 'pdfplumber'
                    }, f, ensure_ascii=False, indent=2)

                has_numbers = True
                numbers_count = len(all_numbers)
                extraction_method = 'pdfplumber'

                # Conta per tipo
                type_counts = {}
                for num in all_numbers:
                    t = num.get('type', 'number')
                    type_counts[t] = type_counts.get(t, 0) + 1

            else:
                # Usa OCR avanzato per PDF rasterizzati
                print(f"PDF {pdf_type} rilevato - Uso OCR avanzato")
                all_numbers, numbers_0deg, numbers_90deg = extract_numbers_advanced(image, min_conf=60)

                # Disegna i rettangoli unificati
                img_with_boxes = draw_unified_boxes(image, numbers_0deg, numbers_90deg)
                page_image = image_to_base64(img_with_boxes)

                # Salva i risultati per la route highlight
                results_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ocr_results.json')
                with open(results_path, 'w', encoding='utf-8') as f:
                    json.dump({
                        'all_numbers': all_numbers,
                        'numbers_0deg': numbers_0deg,
                        'numbers_90deg': numbers_90deg,
                        'extraction_method': 'ocr'
                    }, f, ensure_ascii=False, indent=2)

                has_numbers = True
                numbers_count = len(all_numbers)
                extraction_method = 'ocr'
                type_counts = {
                    '0deg': len(numbers_0deg),
                    '90deg': len(numbers_90deg)
                }

        except Exception as e:
            import traceback
            print(f"Errore nell'estrazione automatica: {traceback.format_exc()}")
            # Fallback: immagine senza rettangoli
            page_image = processor.get_page_image(page_num=0)
            has_numbers = False
            numbers_count = 0
            extraction_method = 'none'
            type_counts = {}
            all_numbers = []

        # Auto-analisi layout per PDF multi-pagina se esiste prompt predefinito
        layout_analysis = None
        auto_layout_executed = False

        if page_count > 1:
            try:
                # Controlla se esiste un prompt layout predefinito
                layout_prompts_file = os.path.join(app.config['LAYOUT_PROMPTS_FOLDER'], 'layout_prompts.json')

                if os.path.exists(layout_prompts_file):
                    with open(layout_prompts_file, 'r', encoding='utf-8') as f:
                        layout_data = json.load(f)

                    # Trova il prompt predefinito
                    default_prompt = None
                    for prompt in layout_data.get('prompts', []):
                        if prompt.get('is_default', False):
                            default_prompt = prompt
                            break

                    if default_prompt:
                        print(f"PDF multi-pagina rilevato ({page_count} pagine) - Esecuzione auto-analisi layout con prompt predefinito: {default_prompt['name']}")

                        # Esegui analisi layout
                        current_provider = ai_manager.get_current_provider()
                        if current_provider:
                            provider_name = ai_manager.get_current_provider_name()

                            # Analizza tutte le pagine
                            results = []
                            for page_num in range(page_count):
                                page_image_b64 = processor.get_page_image(page_num=page_num)

                                try:
                                    analysis = current_provider.analyze_vision(
                                        default_prompt['content'],
                                        page_image_b64
                                    )
                                    results.append({
                                        'page': page_num + 1,
                                        'analysis': analysis
                                    })
                                except Exception as e:
                                    print(f"Errore analisi pagina {page_num + 1}: {str(e)}")
                                    results.append({
                                        'page': page_num + 1,
                                        'error': str(e)
                                    })

                            layout_analysis = {
                                'prompt_name': default_prompt['name'],
                                'prompt_id': default_prompt['id'],
                                'provider': provider_name,
                                'results': results
                            }
                            auto_layout_executed = True
                            print(f"Auto-analisi layout completata con {provider_name}")

            except Exception as e:
                import traceback
                print(f"Errore durante auto-analisi layout: {traceback.format_exc()}")
                # Non bloccare l'upload se l'analisi layout fallisce

        return jsonify({
            'success': True,
            'pdf_type': pdf_type,
            'page_count': page_count,
            'pages_info': pages_info,
            'full_text': full_text,
            'page_image': page_image,
            'has_numbers': has_numbers,
            'numbers': all_numbers,
            'numbers_count': numbers_count,
            'extraction_method': extraction_method,
            'type_counts': type_counts,
            'auto_layout_executed': auto_layout_executed,
            'layout_analysis': layout_analysis
        })

    return jsonify({'error': 'Invalid file type'}), 400


@app.route('/get_page/<int:page_num>')
def get_page(page_num):
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'current.pdf')
    if not os.path.exists(filepath):
        return jsonify({'error': 'No PDF loaded'}), 400

    processor = PDFProcessor(filepath)
    page_image = processor.get_page_image(page_num=page_num)

    return jsonify({
        'success': True,
        'page_image': page_image
    })


@app.route('/extract_pdfplumber', methods=['POST'])
def extract_pdfplumber():
    data = request.json
    page_num = data.get('page_num', 0)
    rotation = data.get('rotation', 0)

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'current.pdf')
    if not os.path.exists(filepath):
        return jsonify({'error': 'No PDF loaded'}), 400

    processor = PDFProcessor(filepath)
    extracted_data = processor.extract_with_pdfplumber(
        page_num=page_num,
        rotation=rotation
    )

    return jsonify({
        'success': True,
        'data': extracted_data,
        'word_count': len(extracted_data)
    })


@app.route('/extract_ocr', methods=['POST'])
def extract_ocr():
    data = request.json
    page_num = data.get('page_num', 0)
    psm_mode = data.get('psm_mode', 6)
    rotation = data.get('rotation', 0)

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'current.pdf')
    if not os.path.exists(filepath):
        return jsonify({'error': 'No PDF loaded'}), 400

    processor = PDFProcessor(filepath)

    # Get page as PIL image
    image = processor.get_page_as_pil(page_num=page_num, dpi=300)
    if not image:
        return jsonify({'error': 'Could not load page'}), 500

    # Apply rotation if needed
    if rotation != 0:
        image = image.rotate(-rotation, expand=True)

    # Run OCR
    custom_config = f'--psm {psm_mode} --oem 3'
    text = pytesseract.image_to_string(image, config=custom_config, lang='ita+eng')

    ocr_data = pytesseract.image_to_data(
        image,
        config=custom_config,
        lang='ita+eng',
        output_type=pytesseract.Output.DICT
    )

    # Format OCR data
    words = []
    if 'text' in ocr_data:
        for i, word in enumerate(ocr_data['text']):
            if word.strip():
                words.append({
                    'text': word,
                    'conf': ocr_data['conf'][i],
                    'left': ocr_data['left'][i],
                    'top': ocr_data['top'][i],
                    'width': ocr_data['width'][i],
                    'height': ocr_data['height'][i]
                })

    return jsonify({
        'success': True,
        'text': text,
        'words': words,
        'word_count': len(words)
    })


@app.route('/extract_numbers_advanced', methods=['POST'])
def extract_numbers_advanced_route():
    """Advanced number extraction with preprocessing and filtering"""
    data = request.json
    page_num = data.get('page_num', 0)
    min_conf = data.get('min_conf', 60)

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'current.pdf')
    if not os.path.exists(filepath):
        return jsonify({'error': 'No PDF loaded'}), 400

    processor = PDFProcessor(filepath)

    # Get page as PIL image
    image = processor.get_page_as_pil(page_num=page_num, dpi=300)
    if not image:
        return jsonify({'error': 'Could not load page'}), 500

    # Salva l'immagine originale per future operazioni
    original_path = os.path.join(app.config['UPLOAD_FOLDER'], 'original.png')
    image.save(original_path)

    # Extract numbers with advanced processing (returns 3 values now)
    all_numbers, numbers_0deg, numbers_90deg = extract_numbers_advanced(image, min_conf=min_conf)

    # Draw unified boxes on image
    img_with_boxes = draw_unified_boxes(image, numbers_0deg, numbers_90deg)
    img_base64 = image_to_base64(img_with_boxes)

    # Salva i risultati per la route highlight
    results_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ocr_results.json')
    with open(results_path, 'w', encoding='utf-8') as f:
        json.dump({
            'all_numbers': all_numbers,
            'numbers_0deg': numbers_0deg,
            'numbers_90deg': numbers_90deg
        }, f, ensure_ascii=False, indent=2)

    return jsonify({
        'success': True,
        'numbers': all_numbers,
        'count': len(all_numbers),
        'count_0deg': len(numbers_0deg),
        'count_90deg': len(numbers_90deg),
        'image': img_base64
    })


@app.route('/highlight/<int:item_id>')
def highlight_item(item_id):
    """Evidenzia un elemento specifico sull'immagine"""
    try:
        # Carica l'immagine originale
        original_path = os.path.join(app.config['UPLOAD_FOLDER'], 'original.png')
        if not os.path.exists(original_path):
            return jsonify({'error': 'Immagine originale non trovata'}), 404

        image = Image.open(original_path)

        # Carica i risultati
        results_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ocr_results.json')
        if not os.path.exists(results_path):
            return jsonify({'error': 'Risultati non trovati'}), 404

        with open(results_path, 'r', encoding='utf-8') as f:
            results = json.load(f)

        extraction_method = results.get('extraction_method', 'ocr')

        if extraction_method == 'pdfplumber':
            # Usa rettangoli colorati per tipo
            all_numbers = results['all_numbers']
            img_with_boxes = draw_pdfplumber_boxes(image, 300, all_numbers, highlight_id=item_id)
        else:
            # Usa rettangoli blu/fucsia per OCR
            numbers_0deg = results.get('numbers_0deg', [])
            numbers_90deg = results.get('numbers_90deg', [])
            img_with_boxes = draw_unified_boxes(image, numbers_0deg, numbers_90deg, highlight_id=item_id)

        img_base64 = image_to_base64(img_with_boxes)

        return jsonify({
            'success': True,
            'image': img_base64
        })

    except Exception as e:
        import traceback
        print(f"Errore highlight: {traceback.format_exc()}")
        return jsonify({'error': f'Errore: {str(e)}'}), 500


@app.route('/opus/analyze', methods=['POST'])
def opus_analyze():
    """Feature 1: Intelligent text analysis with Opus"""
    try:
        # Load extracted data
        results_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ocr_results.json')
        if not os.path.exists(results_path):
            return jsonify({'error': 'Nessun dato estratto disponibile'}), 400

        with open(results_path, 'r', encoding='utf-8') as f:
            results = json.load(f)

        all_numbers = results.get('all_numbers', [])

        # Get PDF type from session or default
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'current.pdf')
        processor = PDFProcessor(filepath)
        pdf_type, _ = processor.detect_pdf_type()

        # Call Opus analysis
        analysis_result = analyze_text_with_opus(all_numbers, pdf_type)

        # Save analysis results for template generation
        if analysis_result.get('success'):
            ai_results_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ai_results.json')
            existing_data = {}
            if os.path.exists(ai_results_path):
                with open(ai_results_path, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)

            existing_data['analysis'] = analysis_result.get('analysis')

            with open(ai_results_path, 'w', encoding='utf-8') as f:
                json.dump(existing_data, f, ensure_ascii=False, indent=2)

        return jsonify(analysis_result)

    except Exception as e:
        import traceback
        print(f"Errore analisi Opus: {traceback.format_exc()}")
        return jsonify({'error': f'Errore: {str(e)}'}), 500


@app.route('/opus/vision', methods=['POST'])
def opus_vision():
    """Feature 2: Vision-guided extraction with Opus"""
    try:
        data = request.json
        image_base64 = data.get('image')

        if not image_base64:
            return jsonify({'error': 'Immagine non fornita'}), 400

        # Get PDF type
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'current.pdf')
        processor = PDFProcessor(filepath)
        pdf_type, _ = processor.detect_pdf_type()

        # Call Opus vision analysis
        vision_result = vision_guided_extraction(image_base64, pdf_type)

        return jsonify(vision_result)

    except Exception as e:
        import traceback
        print(f"Errore visione Opus: {traceback.format_exc()}")
        return jsonify({'error': f'Errore: {str(e)}'}), 500


@app.route('/opus/ask', methods=['POST'])
def opus_ask():
    """Feature 3: Question-Answering with Opus"""
    try:
        data = request.json
        question = data.get('question')

        if not question:
            return jsonify({'error': 'Domanda non fornita'}), 400

        # Load full text
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'current.pdf')
        processor = PDFProcessor(filepath)
        full_text = processor.get_full_text_pdfplumber()

        # Load extracted data
        results_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ocr_results.json')
        extracted_data = []
        if os.path.exists(results_path):
            with open(results_path, 'r', encoding='utf-8') as f:
                results = json.load(f)
                extracted_data = results.get('all_numbers', [])

        # Call Opus Q&A
        qa_result = answer_question_about_pdf(question, full_text, extracted_data)

        return jsonify(qa_result)

    except Exception as e:
        import traceback
        print(f"Errore Q&A Opus: {traceback.format_exc()}")
        return jsonify({'error': f'Errore: {str(e)}'}), 500


@app.route('/opus/summarize', methods=['POST'])
def opus_summarize():
    """Feature 4: Automatic document summarization with Opus"""
    try:
        # Load full text
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'current.pdf')
        processor = PDFProcessor(filepath)
        full_text = processor.get_full_text_pdfplumber()
        pdf_type, _ = processor.detect_pdf_type()

        # Load extracted data
        results_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ocr_results.json')
        extracted_data = []
        if os.path.exists(results_path):
            with open(results_path, 'r', encoding='utf-8') as f:
                results = json.load(f)
                extracted_data = results.get('all_numbers', [])

        # Call Opus summarization
        summary_result = summarize_document(full_text, extracted_data, pdf_type)

        # Save summary results for template generation
        if summary_result.get('success'):
            ai_results_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ai_results.json')
            existing_data = {}
            if os.path.exists(ai_results_path):
                with open(ai_results_path, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)

            existing_data['summary'] = summary_result.get('summary')

            with open(ai_results_path, 'w', encoding='utf-8') as f:
                json.dump(existing_data, f, ensure_ascii=False, indent=2)

        return jsonify(summary_result)

    except Exception as e:
        import traceback
        print(f"Errore riepilogo Opus: {traceback.format_exc()}")
        return jsonify({'error': f'Errore: {str(e)}'}), 500


@app.route('/opus/status')
def opus_status():
    """Check if Opus integration is available"""
    if DEMO_MODE:
        return jsonify({
            'enabled': True,
            'message': '⚠️ DEMO MODE - Risposte simulate'
        })
    return jsonify({
        'enabled': anthropic_client is not None,
        'message': 'Claude Opus integrato' if anthropic_client else 'Configura ANTHROPIC_API_KEY'
    })


# ============================================================================
# AI PROVIDER MANAGEMENT ENDPOINTS
# ============================================================================

@app.route('/ai/providers')
def get_ai_providers():
    """Get list of available AI providers"""
    available = ai_manager.get_available_providers()
    current = ai_manager.current_provider
    return jsonify({
        'providers': available,
        'current': current,
        'current_name': ai_manager.get_current_provider_name()
    })


@app.route('/ai/provider/set', methods=['POST'])
def set_ai_provider():
    """Set the current AI provider"""
    data = request.json
    provider_key = data.get('provider')

    if not provider_key:
        return jsonify({'error': 'Provider key required'}), 400

    success = ai_manager.set_provider(provider_key)
    if success:
        return jsonify({
            'success': True,
            'provider': provider_key,
            'provider_name': ai_manager.get_current_provider_name(),
            'capabilities': ai_manager.get_current_capabilities()
        })
    else:
        return jsonify({'error': 'Provider not available'}), 404


@app.route('/ai/status')
def ai_status():
    """Get current AI provider status"""
    if DEMO_MODE:
        return jsonify({
            'enabled': True,
            'provider': 'demo',
            'message': '⚠️ DEMO MODE - Risposte simulate'
        })

    if ai_manager.is_any_available():
        return jsonify({
            'enabled': True,
            'provider': ai_manager.current_provider,
            'provider_name': ai_manager.get_current_provider_name(),
            'message': f'{ai_manager.get_current_provider_name()} attivo',
            'available_providers': ai_manager.get_available_providers(),
            'capabilities': ai_manager.get_current_capabilities()
        })
    else:
        return jsonify({
            'enabled': False,
            'provider': None,
            'message': 'Nessun provider AI configurato. Aggiungi API keys al file .env'
        })


@app.route('/download_results/<format>', methods=['POST'])
def download_results(format='txt'):
    """Download extraction results in various formats (txt, csv, json)"""
    try:
        # Get data from request body
        request_data = request.get_json()

        if request_data and 'data' in request_data:
            # Use data sent from frontend (currently displayed results)
            all_numbers = request_data['data']
            extraction_method = 'current_view'
        else:
            # Fallback: Load from file if no data sent
            results_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ocr_results.json')
            if not os.path.exists(results_path):
                return jsonify({'error': 'Nessun dato estratto disponibile'}), 404

            with open(results_path, 'r', encoding='utf-8') as f:
                results = json.load(f)

            all_numbers = results.get('all_numbers', [])
            extraction_method = results.get('extraction_method', 'unknown')

        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')

        if format == 'txt':
            # Create formatted text file
            output = io.StringIO()
            output.write(f"PDF Analyzer - Risultati Estrazione\n")
            output.write(f"Data: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            output.write(f"Metodo di estrazione: {extraction_method}\n")
            output.write(f"Elementi totali: {len(all_numbers)}\n")
            output.write("=" * 80 + "\n\n")

            for idx, item in enumerate(all_numbers, 1):
                output.write(f"Elemento {idx}:\n")
                output.write(f"  Testo: {item['text']}\n")
                if 'type' in item:
                    output.write(f"  Tipo: {item['type']}\n")
                if 'confidence' in item:
                    output.write(f"  Confidenza: {item['confidence']}%\n")
                if 'context' in item:
                    output.write(f"  Contesto: {item['context']}\n")
                if 'bbox' in item:
                    bbox = item['bbox']
                    output.write(f"  Posizione: x={bbox.get('x', 0):.1f}, y={bbox.get('y', 0):.1f}, ")
                    output.write(f"w={bbox.get('width', 0):.1f}, h={bbox.get('height', 0):.1f}\n")
                if 'source' in item:
                    output.write(f"  Origine: {item['source']}\n")
                output.write("\n")

            # Convert to bytes
            output_bytes = io.BytesIO(output.getvalue().encode('utf-8'))
            output_bytes.seek(0)

            return send_file(
                output_bytes,
                mimetype='text/plain',
                as_attachment=True,
                download_name=f'pdf_extraction_{timestamp}.txt'
            )

        elif format == 'csv':
            # Create CSV file
            output = io.StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow(['ID', 'Testo', 'Tipo', 'Confidenza', 'Contesto', 'X', 'Y', 'Larghezza', 'Altezza', 'Origine'])

            # Write data
            for idx, item in enumerate(all_numbers, 1):
                bbox = item.get('bbox', {})
                writer.writerow([
                    idx,
                    item.get('text', ''),
                    item.get('type', ''),
                    item.get('confidence', ''),
                    item.get('context', ''),
                    bbox.get('x', ''),
                    bbox.get('y', ''),
                    bbox.get('width', ''),
                    bbox.get('height', ''),
                    item.get('source', '')
                ])

            # Convert to bytes
            output_bytes = io.BytesIO(output.getvalue().encode('utf-8-sig'))  # UTF-8 with BOM for Excel
            output_bytes.seek(0)

            return send_file(
                output_bytes,
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'pdf_extraction_{timestamp}.csv'
            )

        elif format == 'json':
            # Return JSON file with pretty formatting
            output_data = {
                'metadata': {
                    'timestamp': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'extraction_method': extraction_method,
                    'total_items': len(all_numbers)
                },
                'results': all_numbers
            }

            output_bytes = io.BytesIO(json.dumps(output_data, indent=2, ensure_ascii=False).encode('utf-8'))
            output_bytes.seek(0)

            return send_file(
                output_bytes,
                mimetype='application/json',
                as_attachment=True,
                download_name=f'pdf_extraction_{timestamp}.json'
            )

        else:
            return jsonify({'error': 'Formato non supportato. Usa: txt, csv, o json'}), 400

    except Exception as e:
        import traceback
        print(f"Errore download: {traceback.format_exc()}")
        return jsonify({'error': f'Errore durante il download: {str(e)}'}), 500


@app.route('/get_extraction_results', methods=['GET'])
def get_extraction_results():
    """Return OCR/pdfplumber extraction results"""
    try:
        results_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ocr_results.json')
        if not os.path.exists(results_path):
            return jsonify({'success': False, 'error': 'No results available'}), 404

        with open(results_path, 'r', encoding='utf-8') as f:
            results = json.load(f)

        return jsonify({
            'success': True,
            'numbers': results.get('all_numbers', []),
            'extraction_method': results.get('extraction_method', 'unknown')
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/get_pdf_text', methods=['GET'])
def get_pdf_text():
    """Return full PDF text extracted with pdfplumber"""
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'current.pdf')
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'error': 'No PDF loaded'}), 404

        processor = PDFProcessor(filepath)
        full_text = processor.get_full_text_pdfplumber()

        return jsonify({
            'success': True,
            'text': full_text
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/get_ai_results', methods=['GET'])
def get_ai_results():
    """Return stored AI analysis and summary results"""
    try:
        ai_results_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ai_results.json')

        if not os.path.exists(ai_results_path):
            return jsonify({
                'success': True,
                'analysis': None,
                'summary': None
            })

        with open(ai_results_path, 'r', encoding='utf-8') as f:
            ai_results = json.load(f)

        return jsonify({
            'success': True,
            'analysis': ai_results.get('analysis'),
            'summary': ai_results.get('summary')
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def retry_with_increased_temperature(provider, prompt, text, provider_name, temp_increment=0.1):
    """
    Retry AI call with temporarily increased temperature without modifying saved values.

    Args:
        provider: AI provider instance
        prompt: Prompt string
        text: Text to analyze
        provider_name: Provider name string
        temp_increment: Temperature increment (default 0.1)

    Returns:
        Response text or None if retry fails.
    """
    try:
        # Determine provider type and call with increased temperature
        if 'gemini' in provider_name.lower():
            # Gemini retry with temp increment
            import google.generativeai as genai
            new_temp = min(0.0 + temp_increment, 1.0)  # Base 0.0, capped at 1.0
            generation_config = genai.GenerationConfig(
                temperature=new_temp,
                top_p=0.1,
                top_k=1,
                max_output_tokens=8192,
            )
            response = provider.client.generate_content(
                f"{prompt}\n\n{text}",
                generation_config=generation_config
            )
            # Handle safety blocks
            if not response.candidates:
                return None
            candidate = response.candidates[0]
            if candidate.finish_reason and candidate.finish_reason != 1:
                return None
            return response.text

        elif 'claude' in provider_name.lower():
            # Claude retry with temp increment
            model = "claude-opus-4-1-20250805" if "opus" in provider_name.lower() else "claude-sonnet-4-5-20250929"
            new_temp = min(1.0 + temp_increment, 1.0)  # Base 1.0, capped at 1.0 by API
            message = provider.client.messages.create(
                model=model,
                max_tokens=4096,
                temperature=new_temp,
                messages=[{
                    "role": "user",
                    "content": f"{prompt}\n\n{text}"
                }]
            )
            return message.content[0].text

        elif 'gpt' in provider_name.lower() or 'openai' in provider_name.lower():
            # OpenAI retry with temp increment
            new_temp = min(0.7 + temp_increment, 2.0)  # Base 0.7, capped at 2.0
            response = provider.client.chat.completions.create(
                model="gpt-4.1-2025-04-14",
                messages=[
                    {"role": "system", "content": "You are a helpful AI assistant analyzing PDF documents."},
                    {"role": "user", "content": f"{prompt}\n\n{text}"}
                ],
                max_tokens=4096,
                temperature=new_temp
            )
            return response.choices[0].message.content

        elif 'novita' in provider_name.lower() or 'qwen' in provider_name.lower():
            # Novita AI retry with temp increment
            new_temp = min(0.6 + temp_increment, 1.0)  # Base 0.6, capped at 1.0
            response = provider.client.chat.completions.create(
                model="qwen/qwen3-vl-235b-a22b-thinking",
                messages=[
                    {"role": "system", "content": "You are a helpful AI assistant analyzing PDF documents."},
                    {"role": "user", "content": f"{prompt}\n\n{text}"}
                ],
                max_tokens=32768,
                temperature=new_temp,
                top_p=0.95,
                extra_body={
                    "enable_thinking": True,
                    "top_k": 20,
                    "min_p": 0
                }
            )
            return response.choices[0].message.content

        else:
            print(f"Unknown provider type for retry: {provider_name}")
            return None

    except Exception as e:
        print(f"Retry with temperature +{temp_increment} failed: {str(e)}")
        return None


def generate_excel_from_template_with_opus(template_text, extracted_data):
    """
    Use current AI provider to interpret template and generate Excel file
    """
    # Get current AI provider
    current_provider = ai_manager.get_current_provider()

    if not current_provider and not DEMO_MODE:
        return {'error': 'Nessun provider AI configurato'}

    # Prepare data summary
    data_summary = {
        'ocr_numbers': extracted_data.get('ocr_numbers', []),
        'pdfplumber_text_preview': extracted_data.get('pdfplumber_text', '')[:3000],  # First 3000 chars
        'ai_analysis': extracted_data.get('ai_analysis'),
        'ai_summary': extracted_data.get('ai_summary')
    }

    # Add dimensions if available
    dimensions_data = extracted_data.get('dimensions')
    dimensions_provider = None
    if dimensions_data:
        # Check if dimensions contains provider info
        if isinstance(dimensions_data, dict) and 'text' in dimensions_data:
            data_summary['dimensions'] = dimensions_data['text']
            dimensions_provider = dimensions_data.get('provider')
        else:
            data_summary['dimensions'] = dimensions_data

    # Build dimension notice for prompt
    dimension_notice = ""
    dimension_note_hint = ""
    if dimensions_data:
        if dimensions_provider:
            dimension_notice = f"""IMPORTANTE: Sono disponibili dimensioni estratte dal disegno con {dimensions_provider}.
Usa questi dati per popolare i campi di dimensione nel template.
CRITICO: Copia le dimensioni ESATTAMENTE come fornite, senza modificarle o parsarle.
Se le dimensioni contengono valori multipli (es: "1540x1270x835"), mantieni IL VALORE COMPLETO così com'è.
NON estrarre solo il primo numero, usa SEMPRE il valore integrale."""
            dimension_note_hint = f" - Dimensioni estratte con {dimensions_provider}"
        else:
            dimension_notice = """IMPORTANTE: Sono disponibili dimensioni estratte dal disegno.
Usa questi dati per popolare i campi di dimensione nel template.
CRITICO: Copia le dimensioni ESATTAMENTE come fornite, senza modificarle o parsarle.
Se le dimensioni contengono valori multipli (es: "1540x1270x835"), mantieni IL VALORE COMPLETO così com'è.
NON estrarre solo il primo numero, usa SEMPRE il valore integrale."""

    prompt = f"""Hai ricevuto un template per creare un file Excel/CSV e dei dati estratti da un PDF.

TEMPLATE RICHIESTO:
{template_text}

DATI DISPONIBILI:
{json.dumps(data_summary, indent=2, ensure_ascii=False)}

Il tuo compito è:
1. Interpretare il template e capire quale struttura di Excel/CSV viene richiesta
2. Mappare i dati estratti (numeri OCR, testo pdfplumber, analisi AI, dimensioni se presenti) ai campi del template
3. Generare una struttura dati JSON che rappresenti il foglio Excel da creare

{dimension_notice}

Rispondi SOLO con un JSON in questo formato:
{{
  "sheet_name": "Nome del foglio",
  "headers": ["Colonna1", "Colonna2", "Colonna3", ...],
  "rows": [
    ["valore1", "valore2", "valore3", ...],
    ["valore4", "valore5", "valore6", ...],
    ...
  ],
  "notes": "Note opzionali su come sono stati mappati i dati{dimension_note_hint}"
}}

REGOLE IMPORTANTI:
1. Se non riesci a mappare alcuni campi, lascia celle vuote (stringa vuota "") con una nota esplicativa
2. Per le DIMENSIONI: usa il valore COMPLETO senza modifiche (es: se hai "1540x1270x835", scrivi "1540x1270x835" nella cella, NON solo "1540")
3. Mantieni SEMPRE il formato originale dei dati estratti, specialmente per dimensioni e misure"""

    try:
        if DEMO_MODE:
            # Return demo structure
            return {
                'success': True,
                'excel_data': {
                    'sheet_name': 'Dati Estratti (DEMO)',
                    'headers': ['Campo', 'Valore', 'Fonte'],
                    'rows': [
                        ['Template', 'Demo Mode Attivo', 'Sistema'],
                        ['Nota', 'Configura API key per elaborazione reale', 'Sistema']
                    ],
                    'notes': 'Esempio generato in DEMO MODE'
                }
            }

        # Use current provider's analyze_text method with progressive retry on safety error
        response_text = None
        provider_name = ai_manager.get_current_provider_name()

        try:
            response_text = current_provider.analyze_text(prompt, "")
        except Exception as e:
            error_msg = str(e).lower()
            # Check if it's a safety/temperature error
            if any(keyword in error_msg for keyword in ['safety', 'finish_reason', 'blocked', 'recitation']):
                print(f"Safety error detected with {provider_name}, retrying with progressively increased temperature...")
                # Progressive retry with increments: +0.1, +0.2, +0.3, +0.4, +0.5
                for increment in [0.1, 0.2, 0.3, 0.4, 0.5]:
                    print(f"  Attempt with temperature +{increment}...")
                    response_text = retry_with_increased_temperature(current_provider, prompt, "", provider_name, increment)
                    if response_text:
                        print(f"  [OK] Success with temperature +{increment}")
                        break

                if not response_text:
                    print(f"  [FAILED] All retry attempts failed")
                    raise e  # Re-raise original error if all retries failed
            else:
                raise e

        # Try to parse JSON from response
        try:
            # Remove markdown code blocks if present
            if '```json' in response_text:
                response_text = response_text.split('```json')[1].split('```')[0]
            elif '```' in response_text:
                response_text = response_text.split('```')[1].split('```')[0]

            excel_data = json.loads(response_text.strip())
            return {'success': True, 'excel_data': excel_data, 'provider': provider_name}

        except json.JSONDecodeError as e:
            return {'error': f'Invalid JSON response from {provider_name}: {str(e)}'}

    except Exception as e:
        provider_name = ai_manager.get_current_provider_name()
        return {'error': f'Error calling {provider_name}: {str(e)}'}


@app.route('/generate_from_template', methods=['POST'])
def generate_from_template():
    """
    Generate Excel/CSV file from template using Claude Opus
    """
    try:
        data = request.json
        template = data.get('template')
        extracted_data = data.get('data')

        if not template:
            return jsonify({'error': 'Template non fornito'}), 400

        if not extracted_data:
            return jsonify({'error': 'Dati estratti non disponibili'}), 400

        # Use current AI provider to interpret template and generate structure
        result = generate_excel_from_template_with_opus(template, extracted_data)

        if 'error' in result:
            return jsonify({'error': result['error']}), 500

        excel_data = result['excel_data']

        # Create Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = excel_data.get('sheet_name', 'Dati')[:31]  # Excel limit

        # Style for headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Write headers
        headers = excel_data.get('headers', [])
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write rows
        rows = excel_data.get('rows', [])
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

        # Add notes if present
        notes = excel_data.get('notes', '')
        if notes:
            notes_row = len(rows) + 3
            ws.cell(row=notes_row, column=1, value='Note:')
            ws.cell(row=notes_row + 1, column=1, value=notes)
            ws.merge_cells(start_row=notes_row + 1, start_column=1,
                          end_row=notes_row + 1, end_column=len(headers))

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'template_output_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        print(f"Errore generazione da template: {traceback.format_exc()}")
        return jsonify({'error': f'Errore: {str(e)}'}), 500


# ============================================================================
# TEMPLATE LIBRARY ROUTES
# ============================================================================

@app.route('/get_pdf_info', methods=['GET'])
def get_pdf_info():
    """Get PDF information (page count, type, etc.)"""
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'current.pdf')
        if not os.path.exists(filepath):
            return jsonify({'error': 'No PDF loaded'}), 404

        processor = PDFProcessor(filepath)
        pdf_type, _ = processor.detect_pdf_type()
        page_count = processor.get_page_count()

        return jsonify({
            'success': True,
            'page_count': page_count,
            'pdf_type': pdf_type
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/extract_all_pages_pdfplumber', methods=['POST'])
def extract_all_pages_pdfplumber():
    """Extract data from all pages using pdfplumber"""
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'current.pdf')
        if not os.path.exists(filepath):
            return jsonify({'error': 'No PDF loaded'}), 404

        processor = PDFProcessor(filepath)
        page_count = processor.get_page_count()

        all_data = []
        for page_num in range(page_count):
            page_data = extract_numbers_from_pdfplumber(filepath, page_num)
            all_data.extend(page_data)

        # Save to results file
        results_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ocr_results.json')
        existing_data = {}
        if os.path.exists(results_path):
            with open(results_path, 'r', encoding='utf-8') as f:
                existing_data = json.load(f)

        existing_data['all_numbers'] = all_data
        existing_data['extraction_method'] = 'pdfplumber'

        with open(results_path, 'w', encoding='utf-8') as f:
            json.dump(existing_data, f, ensure_ascii=False, indent=2)

        return jsonify({
            'success': True,
            'total_items': len(all_data),
            'pages_processed': page_count
        })

    except Exception as e:
        import traceback
        print(f"Error extracting all pages: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500


@app.route('/get_templates', methods=['GET'])
def get_templates():
    """Get list of saved templates"""
    try:
        templates_file = os.path.join(app.config['TEMPLATES_FOLDER'], 'templates.json')

        if not os.path.exists(templates_file):
            return jsonify({'success': True, 'templates': []})

        with open(templates_file, 'r', encoding='utf-8') as f:
            templates_data = json.load(f)

        return jsonify({'success': True, 'templates': templates_data.get('templates', [])})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/save_template', methods=['POST'])
def save_template():
    """Save a new template"""
    try:
        data = request.json
        template_name = data.get('name')
        template_content = data.get('content')

        if not template_name or not template_content:
            return jsonify({'error': 'Nome e contenuto richiesti'}), 400

        templates_file = os.path.join(app.config['TEMPLATES_FOLDER'], 'templates.json')

        # Load existing templates
        if os.path.exists(templates_file):
            with open(templates_file, 'r', encoding='utf-8') as f:
                templates_data = json.load(f)
        else:
            templates_data = {'templates': []}

        # Check if name already exists
        existing_names = [t['name'] for t in templates_data['templates']]
        if template_name in existing_names:
            return jsonify({'error': f'Template "{template_name}" già esistente'}), 400

        # Create new template
        template_id = str(uuid.uuid4())
        new_template = {
            'id': template_id,
            'name': template_name,
            'content': template_content,
            'created_at': datetime.datetime.now().isoformat()
        }

        templates_data['templates'].append(new_template)

        # Save updated templates
        with open(templates_file, 'w', encoding='utf-8') as f:
            json.dump(templates_data, f, ensure_ascii=False, indent=2)

        return jsonify({
            'success': True,
            'template_id': template_id,
            'message': f'Template "{template_name}" salvato'
        })

    except Exception as e:
        import traceback
        print(f"Errore salvataggio template: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500


@app.route('/get_template/<template_id>', methods=['GET'])
def get_template(template_id):
    """Get a specific template by ID"""
    try:
        templates_file = os.path.join(app.config['TEMPLATES_FOLDER'], 'templates.json')

        if not os.path.exists(templates_file):
            return jsonify({'error': 'Template non trovato'}), 404

        with open(templates_file, 'r', encoding='utf-8') as f:
            templates_data = json.load(f)

        # Find template
        template = None
        for t in templates_data['templates']:
            if t['id'] == template_id:
                template = t
                break

        if not template:
            return jsonify({'error': 'Template non trovato'}), 404

        return jsonify({
            'success': True,
            'template': template
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/delete_template/<template_id>', methods=['DELETE'])
def delete_template(template_id):
    """Delete a template"""
    try:
        templates_file = os.path.join(app.config['TEMPLATES_FOLDER'], 'templates.json')

        if not os.path.exists(templates_file):
            return jsonify({'error': 'Template non trovato'}), 404

        with open(templates_file, 'r', encoding='utf-8') as f:
            templates_data = json.load(f)

        # Find and remove template
        original_count = len(templates_data['templates'])
        templates_data['templates'] = [t for t in templates_data['templates'] if t['id'] != template_id]

        if len(templates_data['templates']) == original_count:
            return jsonify({'error': 'Template non trovato'}), 404

        # Save updated templates
        with open(templates_file, 'w', encoding='utf-8') as f:
            json.dump(templates_data, f, ensure_ascii=False, indent=2)

        return jsonify({
            'success': True,
            'message': 'Template eliminato'
        })

    except Exception as e:
        import traceback
        print(f"Errore eliminazione template: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500


# ============================================================================
# DIMENSION PROMPT LIBRARY ROUTES
# ============================================================================

@app.route('/get_dimension_prompts', methods=['GET'])
def get_dimension_prompts():
    """Get list of saved dimension prompts"""
    try:
        prompts_file = os.path.join(app.config['DIMENSION_PROMPTS_FOLDER'], 'dimension_prompts.json')

        if not os.path.exists(prompts_file):
            return jsonify({'success': True, 'prompts': []})

        with open(prompts_file, 'r', encoding='utf-8') as f:
            prompts_data = json.load(f)

        return jsonify({'success': True, 'prompts': prompts_data.get('prompts', [])})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/save_dimension_prompt', methods=['POST'])
def save_dimension_prompt():
    """Save a new dimension prompt"""
    try:
        data = request.json
        prompt_name = data.get('name')
        prompt_content = data.get('content')

        if not prompt_name or not prompt_content:
            return jsonify({'error': 'Nome e contenuto richiesti'}), 400

        prompts_file = os.path.join(app.config['DIMENSION_PROMPTS_FOLDER'], 'dimension_prompts.json')

        # Load existing prompts
        if os.path.exists(prompts_file):
            with open(prompts_file, 'r', encoding='utf-8') as f:
                prompts_data = json.load(f)
        else:
            prompts_data = {'prompts': []}

        # Check if name already exists
        existing_names = [p['name'] for p in prompts_data['prompts']]
        if prompt_name in existing_names:
            return jsonify({'error': f'Prompt "{prompt_name}" già esistente'}), 400

        # Create new prompt
        prompt_id = str(uuid.uuid4())
        new_prompt = {
            'id': prompt_id,
            'name': prompt_name,
            'content': prompt_content,
            'created_at': datetime.datetime.now().isoformat()
        }

        prompts_data['prompts'].append(new_prompt)

        # Save updated prompts
        with open(prompts_file, 'w', encoding='utf-8') as f:
            json.dump(prompts_data, f, ensure_ascii=False, indent=2)

        return jsonify({
            'success': True,
            'prompt_id': prompt_id,
            'message': f'Prompt "{prompt_name}" salvato'
        })

    except Exception as e:
        import traceback
        print(f"Errore salvataggio prompt: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500


@app.route('/get_dimension_prompt/<prompt_id>', methods=['GET'])
def get_dimension_prompt(prompt_id):
    """Get a specific dimension prompt by ID"""
    try:
        prompts_file = os.path.join(app.config['DIMENSION_PROMPTS_FOLDER'], 'dimension_prompts.json')

        if not os.path.exists(prompts_file):
            return jsonify({'error': 'Prompt non trovato'}), 404

        with open(prompts_file, 'r', encoding='utf-8') as f:
            prompts_data = json.load(f)

        # Find prompt
        prompt = None
        for p in prompts_data['prompts']:
            if p['id'] == prompt_id:
                prompt = p
                break

        if not prompt:
            return jsonify({'error': 'Prompt non trovato'}), 404

        return jsonify({
            'success': True,
            'prompt': prompt
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/delete_dimension_prompt/<prompt_id>', methods=['DELETE'])
def delete_dimension_prompt(prompt_id):
    """Delete a dimension prompt"""
    try:
        prompts_file = os.path.join(app.config['DIMENSION_PROMPTS_FOLDER'], 'dimension_prompts.json')

        if not os.path.exists(prompts_file):
            return jsonify({'error': 'Prompt non trovato'}), 404

        with open(prompts_file, 'r', encoding='utf-8') as f:
            prompts_data = json.load(f)

        # Find and remove prompt
        original_count = len(prompts_data['prompts'])
        prompts_data['prompts'] = [p for p in prompts_data['prompts'] if p['id'] != prompt_id]

        if len(prompts_data['prompts']) == original_count:
            return jsonify({'error': 'Prompt non trovato'}), 404

        # Save updated prompts
        with open(prompts_file, 'w', encoding='utf-8') as f:
            json.dump(prompts_data, f, ensure_ascii=False, indent=2)

        return jsonify({
            'success': True,
            'message': 'Prompt eliminato'
        })

    except Exception as e:
        import traceback
        print(f"Errore eliminazione prompt: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500


# ============================================================================
# LAYOUT PROMPT LIBRARY ROUTES
# ============================================================================

@app.route('/get_layout_prompts', methods=['GET'])
def get_layout_prompts():
    """Get list of saved layout prompts"""
    try:
        prompts_file = os.path.join(app.config['LAYOUT_PROMPTS_FOLDER'], 'layout_prompts.json')

        if not os.path.exists(prompts_file):
            return jsonify({'success': True, 'prompts': []})

        with open(prompts_file, 'r', encoding='utf-8') as f:
            prompts_data = json.load(f)

        return jsonify({'success': True, 'prompts': prompts_data.get('prompts', [])})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/save_layout_prompt', methods=['POST'])
def save_layout_prompt():
    """Save a new layout prompt"""
    try:
        data = request.json
        prompt_name = data.get('name')
        prompt_content = data.get('content')

        if not prompt_name or not prompt_content:
            return jsonify({'error': 'Nome e contenuto richiesti'}), 400

        prompts_file = os.path.join(app.config['LAYOUT_PROMPTS_FOLDER'], 'layout_prompts.json')

        # Load existing prompts
        if os.path.exists(prompts_file):
            with open(prompts_file, 'r', encoding='utf-8') as f:
                prompts_data = json.load(f)
        else:
            prompts_data = {'prompts': []}

        # Check for duplicate names
        existing_names = [p['name'] for p in prompts_data['prompts']]
        if prompt_name in existing_names:
            return jsonify({'error': f'Prompt "{prompt_name}" già esistente'}), 400

        # Create new prompt
        prompt_id = str(uuid.uuid4())
        new_prompt = {
            'id': prompt_id,
            'name': prompt_name,
            'content': prompt_content,
            'created_at': datetime.datetime.now().isoformat()
        }

        prompts_data['prompts'].append(new_prompt)

        # Save to file
        with open(prompts_file, 'w', encoding='utf-8') as f:
            json.dump(prompts_data, f, ensure_ascii=False, indent=2)

        return jsonify({
            'success': True,
            'prompt_id': prompt_id,
            'message': f'Prompt "{prompt_name}" salvato'
        })

    except Exception as e:
        import traceback
        print(f"Errore salvataggio prompt layout: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500


@app.route('/get_layout_prompt/<prompt_id>', methods=['GET'])
def get_layout_prompt(prompt_id):
    """Get a specific layout prompt by ID"""
    try:
        prompts_file = os.path.join(app.config['LAYOUT_PROMPTS_FOLDER'], 'layout_prompts.json')

        if not os.path.exists(prompts_file):
            return jsonify({'error': 'Prompt non trovato'}), 404

        with open(prompts_file, 'r', encoding='utf-8') as f:
            prompts_data = json.load(f)

        # Find prompt by ID
        prompt = next((p for p in prompts_data['prompts'] if p['id'] == prompt_id), None)

        if not prompt:
            return jsonify({'error': 'Prompt non trovato'}), 404

        return jsonify({
            'success': True,
            'prompt': prompt
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/delete_layout_prompt/<prompt_id>', methods=['DELETE'])
def delete_layout_prompt(prompt_id):
    """Delete a layout prompt"""
    try:
        prompts_file = os.path.join(app.config['LAYOUT_PROMPTS_FOLDER'], 'layout_prompts.json')

        if not os.path.exists(prompts_file):
            return jsonify({'error': 'Prompt non trovato'}), 404

        with open(prompts_file, 'r', encoding='utf-8') as f:
            prompts_data = json.load(f)

        # Find and remove prompt
        original_count = len(prompts_data['prompts'])
        prompts_data['prompts'] = [p for p in prompts_data['prompts'] if p['id'] != prompt_id]

        if len(prompts_data['prompts']) == original_count:
            return jsonify({'error': 'Prompt non trovato'}), 404

        # Save updated prompts
        with open(prompts_file, 'w', encoding='utf-8') as f:
            json.dump(prompts_data, f, ensure_ascii=False, indent=2)

        return jsonify({
            'success': True,
            'message': 'Prompt eliminato'
        })

    except Exception as e:
        import traceback
        print(f"Errore eliminazione prompt layout: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500


# ============================================================================
# DEFAULT PROMPT MANAGEMENT (Generic for all prompt types)
# ============================================================================

def get_prompts_file_path(prompt_type):
    """Get the file path for a prompt type"""
    if prompt_type == 'dimension':
        return os.path.join(app.config['DIMENSION_PROMPTS_FOLDER'], 'dimension_prompts.json')
    elif prompt_type == 'layout':
        return os.path.join(app.config['LAYOUT_PROMPTS_FOLDER'], 'layout_prompts.json')
    elif prompt_type == 'template':
        return os.path.join(app.config['TEMPLATES_FOLDER'], 'templates.json')
    else:
        return None

@app.route('/set_default_prompt/<prompt_type>/<prompt_id>', methods=['POST'])
def set_default_prompt(prompt_type, prompt_id):
    """Set a prompt as default for its type"""
    try:
        prompts_file = get_prompts_file_path(prompt_type)
        if not prompts_file:
            return jsonify({'error': 'Tipo prompt non valido'}), 400

        if not os.path.exists(prompts_file):
            return jsonify({'error': 'File prompt non trovato'}), 404

        # Load prompts
        with open(prompts_file, 'r', encoding='utf-8') as f:
            prompts_data = json.load(f)

        # Get key name (templates vs prompts)
        key = 'templates' if prompt_type == 'template' else 'prompts'

        # Find the prompt and update defaults
        prompt_found = False
        for prompt in prompts_data[key]:
            if prompt['id'] == prompt_id:
                prompt['is_default'] = True
                prompt_found = True
            else:
                prompt['is_default'] = False

        if not prompt_found:
            return jsonify({'error': 'Prompt non trovato'}), 404

        # Save updated prompts
        with open(prompts_file, 'w', encoding='utf-8') as f:
            json.dump(prompts_data, f, ensure_ascii=False, indent=2)

        return jsonify({
            'success': True,
            'message': 'Prompt predefinito impostato'
        })

    except Exception as e:
        import traceback
        print(f"Errore impostazione prompt predefinito: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500


@app.route('/remove_default_prompt/<prompt_type>', methods=['POST'])
def remove_default_prompt(prompt_type):
    """Remove default status from all prompts of a type"""
    try:
        prompts_file = get_prompts_file_path(prompt_type)
        if not prompts_file:
            return jsonify({'error': 'Tipo prompt non valido'}), 400

        if not os.path.exists(prompts_file):
            return jsonify({'success': True})  # No file means no defaults

        # Load prompts
        with open(prompts_file, 'r', encoding='utf-8') as f:
            prompts_data = json.load(f)

        # Get key name (templates vs prompts)
        key = 'templates' if prompt_type == 'template' else 'prompts'

        # Remove all defaults
        for prompt in prompts_data[key]:
            prompt['is_default'] = False

        # Save updated prompts
        with open(prompts_file, 'w', encoding='utf-8') as f:
            json.dump(prompts_data, f, ensure_ascii=False, indent=2)

        return jsonify({
            'success': True,
            'message': 'Prompt predefinito rimosso'
        })

    except Exception as e:
        import traceback
        print(f"Errore rimozione prompt predefinito: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500


@app.route('/get_default_prompt/<prompt_type>', methods=['GET'])
def get_default_prompt(prompt_type):
    """Get the default prompt for a type"""
    try:
        prompts_file = get_prompts_file_path(prompt_type)
        if not prompts_file:
            return jsonify({'error': 'Tipo prompt non valido'}), 400

        if not os.path.exists(prompts_file):
            return jsonify({'success': True, 'default_prompt': None})

        # Load prompts
        with open(prompts_file, 'r', encoding='utf-8') as f:
            prompts_data = json.load(f)

        # Get key name (templates vs prompts)
        key = 'templates' if prompt_type == 'template' else 'prompts'

        # Find default prompt
        default_prompt = None
        for prompt in prompts_data[key]:
            if prompt.get('is_default', False):
                default_prompt = prompt
                break

        return jsonify({
            'success': True,
            'default_prompt': default_prompt
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/analyze_layout', methods=['POST'])
def analyze_layout():
    """Analyze entire PDF document to identify pages with drawings and their layout type"""
    try:
        # Use current AI provider
        current_provider = ai_manager.get_current_provider()
        if not current_provider:
            return jsonify({'error': 'Nessun provider AI configurato. Aggiungi le API keys al file .env'}), 400

        data = request.json
        prompt = data.get('prompt')

        if not prompt:
            return jsonify({'error': 'Prompt richiesto'}), 400

        provider_name = ai_manager.get_current_provider_name()

        # Check vision capability
        capabilities = ai_manager.get_current_capabilities()
        if not capabilities.get('vision_analysis'):
            return jsonify({'error': f'Il provider corrente ({provider_name}) non supporta analisi visione'}), 400

        # Get current PDF file (always saved as 'current.pdf')
        pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], 'current.pdf')

        if not os.path.exists(pdf_path):
            return jsonify({'error': 'Nessun PDF caricato'}), 400

        # Open PDF and get page count
        pdf_document = fitz.open(pdf_path)
        total_pages = len(pdf_document)

        # Analyze all pages
        results = []
        for page_num in range(total_pages):
            page = pdf_document[page_num]

            # Convert page to image
            zoom = 2.0  # Higher resolution for better analysis
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)

            # Convert to PNG bytes
            img_data = pix.tobytes("png")

            # Convert to base64
            image_base64 = base64.b64encode(img_data).decode('utf-8')

            # Create page-specific prompt
            page_prompt = f"{prompt}\n\nPagina {page_num + 1} di {total_pages}"

            # Analyze page with AI
            try:
                analysis = current_provider.analyze_vision(page_prompt, image_base64)
                results.append({
                    'page': page_num + 1,
                    'analysis': analysis
                })
            except Exception as page_error:
                results.append({
                    'page': page_num + 1,
                    'analysis': f'Errore analisi: {str(page_error)}'
                })

        pdf_document.close()

        # Format results
        formatted_analysis = f"=== ANALISI LAYOUT DOCUMENTO ===\n"
        formatted_analysis += f"Totale pagine: {total_pages}\n"
        formatted_analysis += f"Provider: {provider_name}\n"
        formatted_analysis += f"\n{'='*60}\n\n"

        for result in results:
            formatted_analysis += f"PAGINA {result['page']}:\n"
            formatted_analysis += f"{result['analysis']}\n"
            formatted_analysis += f"\n{'-'*60}\n\n"

        return jsonify({
            'success': True,
            'analysis': formatted_analysis,
            'pages_analyzed': len(results),
            'total_pages': total_pages,
            'provider': provider_name
        })

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Errore analisi layout: {error_details}")
        provider_name = ai_manager.get_current_provider_name()

        return jsonify({'error': f'Errore durante l\'analisi layout: {str(e)}'}), 500


@app.route('/extract_dimensions', methods=['POST'])
def extract_dimensions():
    """Extract dimensions from PDF page using current AI provider"""
    try:
        # Use current AI provider instead of hardcoded Claude
        current_provider = ai_manager.get_current_provider()
        if not current_provider:
            return jsonify({'error': 'Nessun provider AI configurato. Aggiungi le API keys al file .env'}), 400

        data = request.json
        prompt = data.get('prompt')
        image_base64 = data.get('image')

        if not prompt:
            return jsonify({'error': 'Prompt richiesto'}), 400

        provider_name = ai_manager.get_current_provider_name()

        # Use vision API for all providers (better accuracy for dimension detection)
        if not image_base64:
            return jsonify({'error': 'Immagine richiesta per l\'estrazione dimensioni'}), 400

        capabilities = ai_manager.get_current_capabilities()
        if not capabilities.get('vision_analysis'):
            return jsonify({'error': f'Il provider corrente ({provider_name}) non supporta analisi visione'}), 400

        dimensions_text = current_provider.analyze_vision(prompt, image_base64)

        return jsonify({
            'success': True,
            'dimensions': dimensions_text,
            'provider': provider_name
        })

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Errore estrazione dimensioni: {error_details}")
        provider_name = ai_manager.get_current_provider_name()

        # Provide helpful message for timeout errors
        error_message = str(e)
        if '524' in error_message or 'timeout' in error_message.lower():
            if 'novita' in provider_name.lower():
                return jsonify({
                    'error': f'{provider_name} timeout: Il server ha impiegato troppo tempo a processare l\'immagine. '
                             'Suggerimento: prova con Claude Opus 4.1, Claude Sonnet 4.5, o GPT-4o per immagini complesse.'
                }), 500
            else:
                return jsonify({'error': f'{provider_name} timeout: Richiesta troppo lunga, riprova o usa un altro provider'}), 500

        return jsonify({'error': f'Error calling {provider_name}: {error_message}'}), 500


@app.route('/extract_dimensions_with_context', methods=['POST'])
def extract_dimensions_with_context():
    """Extract dimensions from PDF page using current AI provider with custom prompt and context data"""
    try:
        # Use current AI provider instead of hardcoded Claude
        current_provider = ai_manager.get_current_provider()
        if not current_provider:
            return jsonify({'error': 'Nessun provider AI configurato. Aggiungi le API keys al file .env'}), 400

        data = request.json
        prompt = data.get('prompt')
        image_base64 = data.get('image')
        context = data.get('context', {})

        if not prompt:
            return jsonify({'error': 'Prompt richiesto'}), 400

        provider_name = ai_manager.get_current_provider_name()

        # Build enhanced prompt with context
        enhanced_prompt = prompt + "\n\n"

        # Add PDF text context if available
        if context.get('pdfplumber_text'):
            enhanced_prompt += "CONTESTO - Testo estratto dal PDF:\n"
            enhanced_prompt += context['pdfplumber_text'][:10000]  # Limit to 10k chars
            enhanced_prompt += "\n\n"

        # Add AI analysis context if available
        if context.get('ai_analysis'):
            enhanced_prompt += "CONTESTO - Analisi AI del documento:\n"
            enhanced_prompt += json.dumps(context['ai_analysis'], ensure_ascii=False, indent=2)[:5000]
            enhanced_prompt += "\n\n"

        # Add AI summary context if available
        if context.get('ai_summary'):
            enhanced_prompt += "CONTESTO - Riepilogo AI del documento:\n"
            enhanced_prompt += json.dumps(context['ai_summary'], ensure_ascii=False, indent=2)[:5000]
            enhanced_prompt += "\n\n"

        # Use vision API for all providers (better accuracy for dimension detection)
        if not image_base64:
            return jsonify({'error': 'Immagine richiesta per l\'estrazione dimensioni'}), 400

        capabilities = ai_manager.get_current_capabilities()
        if not capabilities.get('vision_analysis'):
            return jsonify({'error': f'Il provider corrente ({provider_name}) non supporta analisi visione'}), 400

        enhanced_prompt += "Usa il contesto sopra insieme all'immagine per estrarre le dimensioni richieste con maggiore accuratezza."
        dimensions_text = current_provider.analyze_vision(enhanced_prompt, image_base64)

        return jsonify({
            'success': True,
            'dimensions': dimensions_text,
            'provider': provider_name
        })

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Errore estrazione dimensioni con contesto: {error_details}")
        provider_name = ai_manager.get_current_provider_name()

        # Provide helpful message for timeout errors
        error_message = str(e)
        if '524' in error_message or 'timeout' in error_message.lower():
            if 'novita' in provider_name.lower():
                return jsonify({
                    'error': f'{provider_name} timeout: Il server ha impiegato troppo tempo a processare l\'immagine. '
                             'Suggerimento: prova con Claude Opus 4.1, Claude Sonnet 4.5, o GPT-4o per immagini complesse.'
                }), 500
            else:
                return jsonify({'error': f'{provider_name} timeout: Richiesta troppo lunga, riprova o usa un altro provider'}), 500

        return jsonify({'error': f'Error calling {provider_name}: {error_message}'}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5002)
